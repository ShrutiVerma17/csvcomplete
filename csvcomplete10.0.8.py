# -*- coding: utf-8 -*-
"""
Created on Tue Aug  2 10:29:18 2016

@author: Brett Barbaro, Ludovic Autin, Shruti Verma

Must have "HEADERS" in column 1 to mark header row, and "INCLUDE" column with value for every row to be included
10.0.3 can now take xls and xlsx files.
"""

import csv  # for dealing with .csv files
import urllib
import urllib2  # for getting stuff off the web
import os
from HTMLParser import HTMLParser
from Bio import Entrez
from Bio.SeqUtils.ProtParam import ProteinAnalysis
import requests
import time
from Bio import SeqIO
import openpyxl as xl
import ssl

Entrez.email = "bbarbaro@scripps.edu"

# GOAL: input a csv file containing accession numbers
# output: csv file with accession #s, sequences, homologous pdb files, scores/metrics, OPM files
starttime = time.time()
print("hello")


csvpath = """
/Users/mac/Dropbox (Scripps Research)/OLSON/RCSB v Uniprot comparisonB.xlsx
"""
csvpath = csvpath.strip('\n')
print('csvpath = ' + csvpath)
head, tail = os.path.split(csvpath)
model_dir = head + os.sep
csvname, ext = tail.split('.')
pdbpath = model_dir + 'PDB' + os.sep
all_data = []
interactionDict = {'': ''}
interactionDict = {'': '', 'serpina3k': 'Secreted', 'rad23b': 'Cytoplasm;Nucleus;Proteasome', 'pth': 'Secreted', 'ensrnog00000042839': '', 'cyp24a1': 'Mitochondrion', 'ensrnog00000042479': '', 'rsa-14-44': '', 'prkaa1': 'Cytoplasm;Nucleus', 'lrp2': 'Cell membrane;Cell projection;Coated pit;Endosome;Membrane', 'dph1': '', 'cubn': 'Endosome;Lysosome;Membrane', 'kif23': 'Microtubule', 'ranbp9': '', 'gm7589': '', 'prkab1': '', 'slc2a4': 'Cell membrane;Cytoplasm;Membrane', 'kif14': '', 'mapk3': 'Cytoplasm;Membrane;Nucleus', 'rps5': '', 'cyp27b1': 'Membrane;Mitochondrion', 'heparin': '', 'us9': '', 'f2': '', 'zdhhc2': 'Membrane', 'dlg4': 'Cell junction;Cell membrane;Cell projection;Cytoplasm;Membrane;Postsynaptic cell membrane;Synapse', 'gja1': 'Cell junction;Cell membrane;Endoplasmic reticulum;Gap junction;Membrane', 'sftpa1': 'Extracellular matrix;Secreted;Surface film', 'hpx': 'Secreted', 'itm2b': 'Cell membrane;Endosome;Golgi apparatus;Membrane;Secreted', 'ambp': 'Secreted', 'apoh': 'Secreted', 'rpl11': 'Cytoplasm;Nucleus', 'rgd1561333': '', 'kif20b': '', 'kif20a': 'Microtubule', 'sumo3': 'Cytoplasm;Nucleus', 'rpl5l1': '', 'ubc': 'Cytoplasm;Nucleus', 'rhob': 'Cell membrane;Endosome;Membrane;Nucleus', 'rhoa': 'Cell membrane;Cell projection;Cytoplasm;Cytoskeleton;Membrane', 'eef2k': '', 'park7': 'Cell membrane;Cytoplasm;Membrane;Mitochondrion;Nucleus', 'rpl5': 'Cytoplasm;Nucleus', 'akt1': 'Cell membrane;Cytoplasm;Membrane;Nucleus', 'akt2': 'Cell membrane;Cytoplasm;Endosome;Membrane;Nucleus', 'prc1': ''}
newcolumns = 0
# this interactionDict was used for Rat - not appropriate for other species.
# interactionDict = {u'': '', u'Cers3': 'Membrane;Nucleus', u'ENSRNOG00000014551': '', u'Cers1': 'Membrane', u'Man2a2': 'Membrane', u'Ngfg': '', u'Sos1': '', u'Syt1': 'Cell junction;Cytoplasm;Cytoplasmic vesicle;Membrane;Synapse', u'ENSRNOG00000049033': '', u'Gpx1': 'Cytoplasm', u'B2m': 'MHC I;Secreted', u'Gpx4': 'Cytoplasm;Mitochondrion;Nucleus', u'Gpx7': '', u'Gpx8': 'Membrane', u'Flt1': 'Cell membrane;Endosome;Membrane', u'Bcl2': 'Endoplasmic reticulum;Membrane;Mitochondrion;Mitochondrion outer membrane;Nucleus', u'Syf2': 'Nucleus;Spliceosome', u'nme2': '', u'snca': '', u'Uqcrq': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'Akt1': 'Cell membrane;Cytoplasm;Membrane;Nucleus', u'Uqcrh': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'Uqcrb': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'Mlph': '', u'Vps18': '', u'Ezr': 'Cell membrane;Cell projection;Cytoplasm;Cytoskeleton;Membrane', u'Erich2': '', u'Sf3a1': '', u'Sf3a3': '', u'Rab27a': 'Endosome;Lysosome;Membrane', u'Rab27b': 'Membrane', u'Cog2': '', u'March6': 'Membrane', u'Stx1a': 'Cell junction;Cell membrane;Cytoplasmic vesicle;Membrane;Synapse;Synaptosome', u'Scpep1': 'Secreted', u'Mme': 'Cell membrane;Membrane', u'Gm7589': '', u'Adcy4': 'Cell membrane;Cytoplasm;Membrane', u'ENSRNOG00000000938': '', u'Ppap2a': '', u'Adcy3': 'Cell membrane;Cell projection;Cytoplasm;Golgi apparatus;Membrane', u'Kank2': '', u'Sorbs2': 'Cell junction;Cell membrane;Cell projection;Cytoplasm;Membrane', u'Lama1': '', u'Got1': 'Cytoplasm', u'Slc7a6': 'Membrane', u'Slc7a5': 'Cell membrane;Cytoplasm;Membrane', u'Gng11': 'Cell membrane;Membrane', u'Gng12': 'Cell membrane', u'Xiap': 'Cytoplasm;Nucleus', u'Stx12': 'Endosome;Golgi apparatus;Membrane', u'Arhgap11a': '', u'Stx17': 'Cytoplasm;Cytoplasmic vesicle;Endoplasmic reticulum;Membrane', u'Stx16': 'Membrane', u'Gba2': 'Endoplasmic reticulum;Golgi apparatus;Membrane', u'Gba3': '', u'Slc17a7': 'Cell junction;Cytoplasmic vesicle;Membrane;Synapse;Synaptosome', u'Tyms': 'Cytoplasm;Membrane;Mitochondrion;Mitochondrion inner membrane;Nucleus', u'ENSRNOG00000033924': '', u'R3hdml': '', u'Mrpl13': '', u'Psmd4': 'Proteasome', u'ARF1': '', u'Psmd6': 'Proteasome', u'Psmd1': 'Proteasome', u'Psmd3': 'Proteasome', u'Psmd2': 'Proteasome', u'Agpat5': 'Membrane', u'Agpat4': 'Membrane', u'Agpat3': 'Membrane', u'Agpat2': 'Membrane', u'Agpat1': 'Membrane', u'Lcn2': 'Secreted', u'Als2': '', u'Lamtor4': '', u'Lamtor5': '', u'ENSRNOG00000018381': '', u'Lamtor1': 'Cell membrane;Endosome;Lysosome;Membrane', u'Lamtor2': '', u'Lamtor3': 'Endosome;Membrane', u'Ffar1': 'Cell membrane;Membrane', u'Gpr180': 'Membrane', u'Gria3': 'Cell junction;Cell membrane;Membrane;Postsynaptic cell membrane;Synapse', u'Gria1': 'Cell junction;Cell membrane;Cell projection;Endoplasmic reticulum;Endosome;Membrane;Postsynaptic cell membrane;Synapse', u'Gria4': 'Cell junction;Cell membrane;Cell projection;Membrane;Postsynaptic cell membrane;Synapse', u'Cyfip1': 'Cytoplasm', u'Lrrk1': '', u'Lrrk2': '', u'Cyfip2': '', u'rilpl1': '', u'Ahnak': '', u'Rpl5l1': '', u'Emp3': 'Membrane', u'map1a': '', u'Mmp16': 'Cell membrane;Extracellular matrix;Membrane;Secreted', u'ENSRNOG00000046393': '', u'RGD1564854': 'Membrane', u'Cox6c': '', u'RAB21': '', u'Hspd1': 'Mitochondrion', u'Mtor': 'Cytoplasm;Endoplasmic reticulum;Golgi apparatus;Lysosome;Membrane;Mitochondrion;Mitochondrion outer membrane;Nucleus', u'Neu4': '', u'Neu2': 'Cytoplasm', u'Neu3': 'Cell membrane;Membrane', u'FKBP7': '', u'Hnrpa3': '', u'Usp1': 'Nucleus', u'Itsn2': '', u'Ano1': 'Membrane', u'ENSRNOG00000046021': '', u'Eif2b1': '', u'Top2a': 'Nucleus', u'Fars2': 'Mitochondrion', u'Rpl38': '', u'Cstb': 'Cytoplasm', u'Csta': 'Cytoplasm', u'Nme8': 'Cytoplasm', u'ENSRNOG00000022879': '', u'gja1': '', u'Pgrmc2': 'Membrane', u'Fuca1': 'Lysosome', u'Fuca2': 'Secreted', u'ENSRNOG00000003324': '', u'Scn10a': 'Cell membrane;Membrane', u'Farsb': '', u'Rpl3l': '', u'Uap1l1': '', u'Cst5': '', u'Kalrn': 'Cytoplasm;Cytoskeleton', u'itch': '', u'sod1': '', u'Hivep2': 'Nucleus', u'Tgm3': 'Cytoplasm', u'DPYSL2': '', u'Slbp': '', u'ap2b1': '', u'Tor4a': '', u'Pdc': 'Cell projection;Cilium;Cytoplasm;Nucleus', u'Trove2': '', u'Cct8': 'Cytoplasm', u'Cct2': 'Cytoplasm', u'Cct3': 'Cytoplasm', u'Cct7': 'Cytoplasm', u'Cct4': 'Cell projection;Cilium;Cytoplasm;Cytoskeleton', u'Cct5': 'Cytoplasm;Cytoskeleton', u'Lamp2': 'Cell membrane;Cytoplasmic vesicle;Endosome;Lysosome;Membrane', u'RGD1303130': '', u'MEF2D': '', u'mvk': '', u'Cd274': 'Membrane', u'PRKAR2B': '', u'Ppp3ca': 'Cell membrane;Cell projection;Cytoplasm;Membrane', u'Atp12a': 'Membrane', u'Sfxn3': 'Membrane;Mitochondrion', u'Gmpr': '', u'Vamp1': 'Cell junction;Cytoplasmic vesicle;Membrane;Mitochondrion;Mitochondrion outer membrane;Synapse;Synaptosome', u'Vamp3': 'Cell junction;Membrane;Synapse;Synaptosome', u'Vamp2': 'Cell junction;Cell membrane;Cytoplasmic vesicle;Membrane;Synapse;Synaptosome', u'Clns1a': 'Cytoplasm;Cytoskeleton;Nucleus', u'Col7a1': '', u'Vamp7': 'Cell junction;Cytoplasmic vesicle;Endoplasmic reticulum;Endosome;Golgi apparatus;Lysosome;Membrane;Synapse;Synaptosome', u'Vamp8': 'Cell membrane;Endosome;Lysosome;Membrane', u'Usp12': '', u'Glt25d2': '', u'Glt25d1': '', u'Usp14': '', u'Wnt5a': 'Extracellular matrix;Secreted', u'Kcnd3': 'Cell membrane;Cell projection;Membrane', u'Set': 'Cytoplasm;Endoplasmic reticulum;Nucleus', u'Hk2': '', u'Ndufb11': 'Membrane', u'Ndufb10': '', u'Ganc': '', u'RGD1562997': '', u'Myl12a': '', u'H2afx': 'Chromosome;Nucleosome core', u'H2afz': 'Chromosome;Nucleosome core;Nucleus', u'ENSRNOG00000033045': '', u'Mcm5': 'Nucleus', u'Rpl7a': '', u'Sgsh': '', u'Atp5g1': '', u'Cox7c': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'Atp5g3': '', u'Atp5g2': '', u'FYN': '', u'Syp': 'Cell junction;Cytoplasmic vesicle;Membrane;Synapse;Synaptosome', u'TOR1A': '', u'Akr1b7': 'Cytoplasm', u'Hk1': 'Membrane;Mitochondrion;Mitochondrion outer membrane', u'Gabrg2': 'Cell junction;Cell membrane;Cell projection;Cytoplasmic vesicle;Membrane;Postsynaptic cell membrane;Synapse', u'Rab7a': 'Cytoplasmic vesicle;Endosome;Lipid droplet;Lysosome;Membrane', u'Muc16': '', u'ENSRNOG00000013874': '', u'Man2b2': '', u'Man2b1': '', u'Cd44': 'Cell membrane;Cell projection;Membrane', u'Ppp2r5d': '', u'Ppp2r5e': '', u'Ppp2r5a': '', u'Ppp2r5c': '', u'Gpi': 'Cytoplasm;Secreted', u'Hkdc1': '', u'ENSRNOG00000045720': '', u'Wars2': '', u'Chrd': '', u'Csnk2a2': '', u'Pip5k1c': 'Cell junction;Cell membrane;Cell projection;Cytoplasm;Membrane', u'Bid': 'Cytoplasm;Membrane;Mitochondrion', u'Mtp2': '', u'Rnf4': 'Cytoplasm;Nucleus', u'Net1': '', u'Nit2': 'Cytoplasm', u'Sclt1': 'Cytoplasm;Cytoskeleton', u'Ptk2': 'Cell junction;Cell membrane;Cytoplasm;Cytoskeleton;Membrane;Nucleus', u'Cd2ap': 'Cell junction;Cell projection;Cytoplasm;Cytoskeleton', u'Pygb': '', u'Gp9': 'Membrane', u'Mark2': 'Cell membrane;Cell projection;Cytoplasm;Cytoskeleton;Membrane', u'Pygl': '', u'Pygm': '', u'Pfas': '', u'Tubb6': 'Cytoplasm;Cytoskeleton', u'Tubb5': 'Cytoplasm;Cytoskeleton;Microtubule', u'Tubb3': 'Cytoplasm;Cytoskeleton;Microtubule', u'Tubb1': 'Cytoplasm;Cytoskeleton', u'ENSRNOG00000021395': '', u'Tap1': 'Endoplasmic reticulum;Membrane', u'Adcy5': 'Cell membrane;Cell projection;Cilium;Membrane', u'Dgat1': 'Endoplasmic reticulum;Membrane', u'Tbc1d1': '', u'Gsk3a': '', u'Myo5c': '', u'Prkab2': '', u'Sdc1': 'Membrane;Secreted', u'Sdc2': 'Membrane', u'Sdc4': 'Membrane;Secreted', u'Clgn': 'Endoplasmic reticulum', u'Itgav': 'Membrane', u'Gng7': 'Cell membrane;Membrane', u'Gng4': 'Cell membrane', u'Gng2': 'Cell membrane', u'Itgam': 'Membrane', u'Mapk12': 'Cytoplasm;Mitochondrion;Nucleus', u'stxbp5l': '', u'Gng8': 'Cell membrane;Membrane', u'Ctxn3': 'Membrane', u'Ap3m1': 'Cytoplasmic vesicle;Golgi apparatus;Membrane', u'Telo2': '', u'Mgea5': '', u'gnai3': '', u'Ralgds': 'Cytoplasm;Nucleus', u'Itga1': 'Membrane', u'Itga2': 'Membrane', u'Itga3': 'Membrane', u'Cdk1': 'Cytoplasm;Cytoskeleton;Mitochondrion;Nucleus', u'Cdk5': 'Cell junction;Cell membrane;Cell projection;Cytoplasm;Membrane;Nucleus;Postsynaptic cell membrane;Synapse', u'ENSRNOG00000031564': '', u'Trip12': 'Nucleus', u'Apoa1': 'HDL;Secreted', u'Cr2': 'Membrane', u'Wasf1': 'Cell junction;Cytoplasm;Cytoskeleton;Synapse', u'Psmc1': 'Cytoplasm;Membrane;Nucleus;Proteasome', u'Dnm1l': 'Cell junction;Coated pit;Cytoplasm;Cytoplasmic vesicle;Golgi apparatus;Membrane;Mitochondrion;Mitochondrion outer membrane;Peroxisome;Synapse', u'grm7': '', u'Sqstm1': 'Cytoplasm;Cytoplasmic vesicle;Endoplasmic reticulum;Endosome;Lysosome;Nucleus', u'Suv39h1': '', u'Suv39h2': 'Chromosome', u'ENSRNOG00000050788': '', u'Ap2m1': 'Cell membrane;Coated pit;Membrane', u'Prkcsh': '', u'Calr': 'Endoplasmic reticulum;Sarcoplasmic reticulum', u'M6pr': 'Lysosome;Membrane', u'SRC': '', u'Crp': 'Secreted', u'Ptges3l1': '', u'EBI-7705941': '', u'Ap1s1': 'Membrane', u'Pdgfrb': 'Cell membrane;Cytoplasmic vesicle;Lysosome;Membrane', u'Pdgfra': 'Cell membrane;Membrane', u'dynll1': '', u'Dnajc5b': 'Membrane', u'ADRB2': '', u'Wasl': 'Cytoplasm;Cytoskeleton;Nucleus', u'Prkar1b': 'Cell membrane;Membrane', u'gne': '', u'Cs': 'Mitochondrion', u'Cp': 'Secreted', u'OCRL': '', u'Nr3c1': 'Cytoplasm;Cytoskeleton;Mitochondrion;Nucleus', u'ENSRNOG00000015290': '', u'Txlnb': '', u'Abcg5': 'Cell membrane;Membrane', u'Amfr': 'Membrane', u'Kif23': 'Microtubule', u'RGS4': '', u'Dstn': '', u'RGS7': '', u'RGS1': '', u'Arhgap5': '', u'RGS8': '', u'Rassf9': 'Endosome', u'Nras': 'Cell membrane;Golgi apparatus;Membrane', u'Dnm1': 'Cytoplasm;Cytoskeleton;Golgi apparatus;Microtubule', u'Dnm2': 'Cell junction;Cell membrane;Cell projection;Coated pit;Cytoplasm;Cytoplasmic vesicle;Cytoskeleton;Membrane;Microtubule;Postsynaptic cell membrane;Synapse', u'Sphk2': '', u'Sphk1': 'Cell membrane;Cytoplasm;Membrane;Nucleus', u'Ramp1': 'Membrane', u'Sparcl1': 'Extracellular matrix;Secreted', u'Ramp3': 'Cell membrane;Membrane', u'Ramp2': 'Membrane', u'Dvl2': '', u'Dvl3': '', u'Syt11': 'Cell junction;Cytoplasmic vesicle;Membrane;Synapse', u'Bdnf': 'Secreted', u'Zcchc4': '', u'ENSRNOG00000012086': '', u'Atic': '', u'O3far1': '', u'Man2c1': 'Cytoplasm', u'Cd34': 'Membrane', u'Gak': 'Cell junction;Cytoplasm;Golgi apparatus', u'Gaa': 'Lysosome;Membrane', u'Ndufab1': '', u'gabarap': '', u'Cdh1': 'Cell junction;Cell membrane;Endosome;Golgi apparatus;Membrane', u'Mylpf': '', u'Isy1-rab43': '', u'Igsf8': 'Membrane', u'St5': '', u'sptan1': '', u'Supt3h': '', u'Atp5h': 'CF(0);Membrane;Mitochondrion;Mitochondrion inner membrane', u'Atp5o': '', u'Nphs1': 'Cell membrane;Membrane', u'Aldh16a1': '', u'Atp5b': '', u'Atp5d': '', u'Bcap31': 'Membrane', u'Mycn': 'Nucleus', u'Synrg': 'Cytoplasm;Golgi apparatus;Membrane', u'Pkia': '', u'Txlna': '', u'Rpl15': 'Membrane', u'Rpl10': '', u'Rpl11': 'Cytoplasm;Nucleus', u'Timm8b': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'Tuba8': 'Cytoplasm;Cytoskeleton;Microtubule', u'Rragc': '', u'LRRK2': '', u'Ogfod2': '', u'Cubn': 'Endosome;Lysosome;Membrane', u'Xkr8': 'Membrane', u'Lep': 'Secreted', u'Plk1': 'Centromere;Chromosome;Cytoplasm;Cytoskeleton;Kinetochore;Nucleus', u'Clta': 'Coated pit;Cytoplasm;Cytoplasmic vesicle;Cytoskeleton;Membrane', u'Ift27': '', u'Smpd3': 'Cell membrane;Golgi apparatus;Membrane', u'hdac6': 'Nucleus', u'Tctn2': 'Cell projection;Cytoplasm;Cytoskeleton;Membrane', u'Nisch': 'Cell membrane;Cytoplasm;Endosome;Membrane', u'Cltb': 'Coated pit;Cytoplasmic vesicle;Membrane', u'ANKRD27': '', u'Cdc42ep4': '', u'Gusb': 'Lysosome', u'heparin': '', u'Palm2': '', u'Wnt3': 'Extracellular matrix', u'Hk3': '', u'Wnt1': 'Extracellular matrix', u'ldbi': '', u'Wnt4': 'Extracellular matrix', u'SNCA': '', u'ENSRNOG00000045558': '', u'Psmd7': 'Proteasome', u'SOST': '', u'Sod2': 'Mitochondrion', u'Sod1': 'Cytoplasm;Nucleus', u'Ftsj3': 'Nucleus', u'Rock2': 'Cell membrane;Cytoplasm;Cytoskeleton;Membrane;Nucleus', u'pak1': '', u'pak2': '', u'fxyd3': '', u'Pard6g': '', u'Pard6a': 'Cell junction;Cell membrane;Cytoplasm;Cytoskeleton;Membrane;Tight junction', u'Thop1': 'Cytoplasm', u'Prkar2a': 'Cell membrane;Cytoplasm;Membrane', u'Insr': 'Cell membrane;Membrane', u'Prkar2b': 'Cell membrane;Cytoplasm;Membrane', u'Ncl': 'Cytoplasm;Nucleus', u'Uqcrc2': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'Tmed10': 'Cell membrane;Cytoplasmic vesicle;Endoplasmic reticulum;Golgi apparatus;Membrane', u'Gldc': 'Mitochondrion', u'H6pd': '', u'Camk2a': 'Cell junction;Cell membrane;Cell projection;Membrane;Postsynaptic cell membrane;Synapse', u'GGA2': '', u'Uqcrfs1': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'Ampd3': '', u'Tubb4b': 'Cytoplasm;Cytoskeleton;Microtubule', u'Ampd1': '', u'Ghrl': 'Secreted', u'Ins1': 'Secreted', u'Ins2': 'Secreted', u'PRKACA': '', u'slc12a3': '', u'Igf2r': 'Membrane', u'Dhx8': '', u'Atp1a1': 'Cell membrane;Membrane', u'Pgm1': 'Cytoplasm', u'Cul5': '', u'Nrp1': 'Membrane', u'Wnt3a': 'Extracellular matrix', u'Npl': 'Cytoplasm', u'Nlgn1': 'Cell junction;Cell membrane;Membrane;Postsynaptic cell membrane;Synapse', u'Slc9a3r2': 'Cell membrane;Membrane;Nucleus', u'Kif3b': 'Microtubule', u'Nps': 'Secreted', u'Gucy1b1': 'Cytoplasm', u'sumo3': '', u'mllt10': '', u'Usp38': '', u'Doc2b': 'Cell membrane;Cytoplasm;Membrane', u'Trpv5': 'Cell membrane;Membrane', u'Usp35': '', u'Xpo1': 'Cytoplasm;Nucleus', u'Doc2g': '', u'LOC681754': '', u'Pxn': 'Cell junction;Cytoplasm;Cytoskeleton', u'Atp4a': 'Cell membrane;Membrane', u'Tmem208': 'Membrane', u'Tst': 'Mitochondrion', u'Sept9': 'Cytoplasm;Cytoskeleton', u'Sept5': 'Cytoplasm;Cytoskeleton', u'Sept7': 'Cell projection;Centromere;Chromosome;Cilium;Cytoplasm;Cytoskeleton;Flagellum;Kinetochore', u'Cyth2': 'Cell junction;Cell membrane;Cell projection;Cytoplasm;Membrane;Tight junction', u'Sept1': 'Cytoplasm;Cytoskeleton', u'Prkag3': '', u'Sept3': 'Cell junction;Cytoplasm;Cytoskeleton;Synapse', u'Sept2': 'Cell membrane;Cell projection;Cilium;Cytoplasm;Cytoskeleton;Flagellum;Membrane', u'ENSRNOG00000043499': '', u'pld1': '', u'pld2': '', u'Nap1l1': 'Nucleus', u'Plrg1': 'Nucleus;Spliceosome', u'Nap1l4': 'Cytoplasm;Nucleus', u'Tiam1': '', u'Cit': '', u'Tat': '', u'Eps15': '', u'ADRA1B': '', u'Gria2': 'Cell junction;Cell membrane;Endoplasmic reticulum;Membrane;Postsynaptic cell membrane;Synapse', u'Ptprc': 'Membrane', u'Ipo9': 'Nucleus', u'Aacs': 'Cytoplasm', u'Ptprn': 'Cell junction;Cell membrane;Cell projection;Cytoplasmic vesicle;Endosome;Membrane;Nucleus;Synapse', u'Pten': '', u'Ren': '', u'Prmt1': 'Cytoplasm;Nucleus', u'bcl2l11': '', u'ENSRNOG00000046723': '', u'Ppp1r9b': 'Cell junction;Cell membrane;Cell projection;Cytoplasm;Cytoskeleton;Membrane;Nucleus;Synapse', u'Cps1': 'Mitochondrion;Nucleus', u'Agl': '', u'Calm1': 'Cytoplasm;Cytoskeleton', u'Map3k3': '', u'Scg5': 'Secreted', u'Lrrc15': 'Membrane', u'Scg3': 'Cytoplasmic vesicle;Membrane;Secreted', u'Scg2': 'Secreted', u'Aga': 'Lysosome', u'Mtrr': 'Cytoplasm', u'Tubgcp2': 'Cytoplasm;Cytoskeleton', u'Xylb': '', u'Pdcl2': '', u'Pdcl3': 'Cytoplasm', u'Cox5b': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'RT1-Db1': 'Membrane;MHC II', u'RT1-Db2': 'Membrane;MHC II', u'Pak4': '', u'Calca': 'Secreted', u'Pak1': 'Cell junction;Cell membrane;Cell projection;Cytoplasm;Membrane', u'Pak2': 'Cytoplasm;Membrane;Nucleus', u'Pak3': 'Cytoplasm', u'Pcsk6': '', u'hsp90b1': '', u'Atp5f1b': 'CF(1);Membrane;Mitochondrion;Mitochondrion inner membrane', u'Uqcrbl': '', u'Ruvbl1': 'Nucleus', u'Ogt': 'Cell membrane;Cytoplasm;Membrane;Nucleus', u'Kras': 'Cell membrane;Cytoplasm;Membrane', u'mast3_proex': '', u'Braf': '', u'vamp2': 'Membrane', u'Eif2s1': 'Cytoplasm', u'Creb1': 'Nucleus', u'Gast': 'Secreted', u'Lewi': '', u'Palld': 'Cell junction;Cell projection;Cytoplasm;Cytoskeleton', u'tfrc': '', u'Ppt1': 'Lysosome;Secreted', u'Ppt2': 'Lysosome', u'Ehmt1': 'Chromosome', u'Glb1l2': '', u'Ehmt2': 'Chromosome', u'Snap47': 'Cytoplasm;Membrane', u'Efhd2': 'Membrane', u'Pyy': 'Secreted', u'Amz1': '', u'Gnao1': 'Cell membrane;Membrane', u'Tpm4': 'Cytoplasm;Cytoskeleton', u'Tpm1': 'Cytoplasm;Cytoskeleton', u',5': '', u'Ppp1r12a': 'Cytoplasm', u'Ppp1r12b': '', u'Ppp1r12c': '', u'Mmp7': 'Extracellular matrix;Secreted', u'Rab8b': 'Cell membrane;Cytoplasmic vesicle;Membrane', u'Rab8a': 'Cell membrane;Cell projection;Cilium;Cytoplasm;Cytoplasmic vesicle;Cytoskeleton;Endosome;Golgi apparatus;Membrane', u'Rap1gap': '', u'LOC316124': '', u'ENSRNOG00000005581': '', u'Ap2a1': 'Coated pit', u'Ap2a2': 'Cell membrane;Coated pit;Membrane', u'Dnaja3': '', u'Trio': 'Cytoplasm', u'Dnaja1': 'Cytoplasm;Endoplasmic reticulum;Membrane;Microsome;Mitochondrion;Nucleus', u'Klc4': 'Cytoplasm;Cytoskeleton;Microtubule', u'Dnmt3b': '', u'Ngf': 'Secreted', u'Igf1': 'Secreted', u'RGD1308564': '', u'Chm': 'Cytoplasm', u'Apoh': 'Secreted', u'Mrpl51': '', u'Mrpl52': '', u'Acly': 'Cytoplasm', u'Rims1': 'Cell junction;Cell membrane;Membrane;Synapse', u'Pld4': 'Membrane', u'Lactb2': 'Mitochondrion', u'Pld1': 'Cytoplasm;Endoplasmic reticulum;Endosome;Golgi apparatus;Membrane', u'Pld2': 'Membrane', u'Tbp': '', u'Lalba': 'Secreted', u'Mgp': 'Secreted', u'Atp6v0d2': '', u'Atp6v0d1': '', u'Gpr119': 'Cell membrane;Membrane', u'abcg1': '', u'Plcb3': 'Cytoplasm;Membrane', u'Ctsl1': '', u'ENSRNOG00000048608': '', u'Polr2b': 'DNA-directed RNA polymerase', u'ENSRNOG00000048606': '', u'Dennd1a': '', u'Wnt2b': 'Extracellular matrix', u'Hibch': 'Mitochondrion', u'Rbm28': '', u'Gcg': 'Secreted', u'Gck': 'Cytoplasm;Nucleus', u'Wisp1': 'Secreted', u'Ykt6': 'Cytoplasm;Cytoplasmic vesicle;Golgi apparatus;Membrane', u'Slc25a5': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'Gucy1a1': 'Cytoplasm', u'Cd19': 'Membrane', u'Stmn4': 'Cell projection;Golgi apparatus', u'Hsd17b10': 'Mitochondrion', u'sik3': '', u'Rab2a': 'Endoplasmic reticulum;Golgi apparatus;Membrane', u'Eif2ak4': 'Cytoplasm', u'Man2a1': 'Golgi apparatus;Membrane', u'Camkk1': 'Cytoplasm;Nucleus', u'Cacul1': '', u'Gcgr': 'Cell membrane;Membrane', u'Vasp': 'Cell projection', u'Xab2': 'Nucleus;Spliceosome', u'ocrl': '', u'Tp53': 'Cytoplasm;Endoplasmic reticulum;Mitochondrion;Nucleus', u'Cspg5': 'Cell membrane;Endoplasmic reticulum;Golgi apparatus;Membrane', u'F12': 'Secreted', u'F11': '', u'Supt5h': 'Nucleus', u'Lpl': 'Cell membrane;Chylomicron;Membrane;Secreted;VLDL', u'ENSRNOG00000012339': '', u'Rpl31-ps8': '', u'Abcb9': 'Lysosome;Membrane', u'SET': '', u'Zdhhc13': 'Membrane', u'Arpc3': 'Cytoplasm;Cytoskeleton', u'Arpc1b': 'Cytoplasm;Cytoskeleton', u'Pdia4': 'Endoplasmic reticulum', u'Pdia3': 'Endoplasmic reticulum', u'Pdia2': '', u'Arpc5': 'Cell projection;Cytoplasm;Cytoskeleton', u'Arpc4': 'Cytoplasm;Cytoskeleton', u'Dcp1a': '', u'Myh10': 'Cell projection', u'Arse': '', u'Arsg': 'Lysosome', u'Myh14': '', u'Arsa': '', u'Arsb': 'Lysosome', u'ncald': '', u'Arsi': 'Endoplasmic reticulum;Secreted', u'Arsk': 'Secreted', u'Nufip1': 'Nucleus', u'Paqr8': 'Membrane', u'RGD1565767': '', u'Umps': '', u'Gm10136': '', u'Paqr7': 'Membrane', u'Btd': 'Secreted', u'Ank2': '', u'Angptl4': 'Extracellular matrix;Secreted', u'OGT': '', u'Bud31': 'Nucleus;Spliceosome', u'Ankk1': '', u'Uqcrc1': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'Cul1': '', u'Cul2': '', u'Cul3': 'Cell projection;Cilium;Cytoplasm;Flagellum;Golgi apparatus;Nucleus', u'Trappc1': 'Endoplasmic reticulum;Golgi apparatus', u'Trappc3': 'Endoplasmic reticulum;Golgi apparatus', u'Vcl': 'Cell junction;Cell membrane;Cytoplasm;Cytoskeleton;Membrane', u'Mylk3': 'Cytoplasm', u'Vcp': 'Cytoplasm;Endoplasmic reticulum;Nucleus', u'Hepacam2': 'Membrane', u'Mboat4': 'Endoplasmic reticulum;Membrane', u'Igf2bp2': '', u'Dcbld2': 'Membrane', u'RGD1566120': '', u'Phb': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'Cdh9': 'Cell membrane', u'Nsd1': 'Nucleus', u'Cdh2': 'Cell junction;Cell membrane;Membrane', u'Hyal3': 'Cell membrane;Cytoplasmic vesicle;Endoplasmic reticulum;Endosome;Membrane;Secreted', u'Hyal1': 'Lysosome;Secreted', u'Hyal4': '', u'Akt1s1': '', u'Sptbn2': 'Cytoplasm;Cytoskeleton', u'Sptbn1': 'Cytoplasm;Cytoskeleton', u'Rilp': '', u'Eea1': '', u'Eno-2': '', u'Glrx3': 'Cytoplasm', u'cfl1': '', u'Sez6l2': 'Membrane', u'Ostm1': 'Membrane', u'Spdl1': 'Centromere;Chromosome;Cytoplasm;Cytoskeleton;Kinetochore;Nucleus', u'Tktl2': '', u'Polr1b': 'DNA-directed RNA polymerase;Nucleus', u'Txn1': '', u'Rab5a': 'Cell membrane;Cell projection;Cytoplasm;Cytoplasmic vesicle;Endosome;Membrane', u'Rab5c': '', u'Rab5b': '', u'Pcsk2': 'Cytoplasmic vesicle;Secreted', u'Atp5l': 'CF(0);Membrane;Mitochondrion;Mitochondrion inner membrane', u'Pcsk1': 'Cytoplasmic vesicle', u'Fyn': 'Cell membrane;Membrane', u'Pcsk7': 'Golgi apparatus;Membrane', u'Pcsk4': 'Cytoplasmic vesicle;Membrane', u'Pcsk5': 'Membrane;Secreted', u'Rgs4': '', u'Glipr1l2': 'Membrane', u'Cym': '', u'Anpep': 'Cell membrane;Membrane', u'Dpyd': 'Cytoplasm', u'SH3GLB2': '', u'ENSRNOG00000036701': '', u'Mrpl47': '', u'Kif11': '', u'Psma4': 'Cytoplasm;Nucleus;Proteasome', u'Mrpl44': '', u'Pcdh20': 'Membrane', u'Ns5atp9': '', u'Cdc25c': '', u'Cdc25b': 'Cytoplasm;Cytoskeleton', u'Rpl18a': '', u'Rps15a': '', u'Pdk1': 'Mitochondrion', u'Pdk2': 'Mitochondrion', u'Chga': 'Cytoplasmic vesicle;Membrane;Secreted', u'Tubb4a': 'Cytoplasm;Cytoskeleton', u'Wnt11': 'Extracellular matrix', u'Atp5c1': '', u'Grpel1': 'Mitochondrion', u'Grpel2': 'Mitochondrion', u'Smpdl3a': 'Secreted', u'Kcnj5': 'Membrane', u'Kcnj8': 'Membrane', u'Acsbg1': 'Cytoplasm;Cytoplasmic vesicle;Endoplasmic reticulum;Microsome', u'Acsbg2': 'Cytoplasm;Membrane', u'Lgals3': 'Cytoplasm;Nucleus;Secreted;Spliceosome', u'rCG_38845': '', u'RGD1564719': '', u'H2afj': 'Chromosome;Nucleosome core;Nucleus', u'Nrd1': '', u'Gnat1': '', u'Gnat3': 'Cytoplasm', u'Gnat2': '', u'Csnk2b': '', u'A1bg': 'Secreted', u'Nedd8': 'Nucleus', u'Ppp2r1a': '', u'Rnf139': 'Membrane', u'Arnt': 'Nucleus', u'Rtdr1': '', u'arf6': '', u'mib1': '', u'Stxbp1': 'Cytoplasm;Membrane', u'Stxbp2': '', u'Fmr1': 'Cell junction;Cell membrane;Cell projection;Centromere;Chromosome;Cytoplasm;Membrane;Nucleus;Postsynaptic cell membrane;Synapse;Synaptosome', u'Ywhaz': 'Cytoplasm', u'Manba': 'Lysosome', u'GNAS': '', u'Ywhaq': 'Cytoplasm', u'Nt5e': 'Cell membrane;Membrane', u'Ywhah': 'Cytoplasm', u'Ywhab': 'Cytoplasm', u'Ywhag': 'Cytoplasm', u'Ywhae': 'Cytoplasm;Nucleus', u'ywhaz': '', u'Rpl35': '', u'Tmem106b': 'Endosome;Lysosome;Membrane', u'RGD1564883': '', u'Chit1': '', u'Egln1': 'Cytoplasm;Nucleus', u'Rps3a': 'Cytoplasm;Nucleus', u'Myo9b': 'Cytoplasm;Cytoskeleton', u'Etfa': 'Mitochondrion', u'Kif18a': 'Microtubule', u'Pdia6': 'Cell membrane;Endoplasmic reticulum;Membrane', u'Sgms2': 'Cell membrane;Golgi apparatus;Membrane', u'Vdac2': 'Membrane;Mitochondrion;Mitochondrion outer membrane', u'Rcn2': 'Endoplasmic reticulum', u'Gba': '', u'Lrrc30': '', u'Npepl1': '', u'Mttp': '', u'Abcc8': 'Cell membrane;Membrane', u'Abcc9': 'Membrane', u'Prkd3': 'Cytoplasm', u'Bet1': 'Endoplasmic reticulum;Golgi apparatus;Membrane', u'Vars': '', u'Rabgef1': '', u'Kdr': 'Cell junction;Cell membrane;Cytoplasm;Cytoplasmic vesicle;Endoplasmic reticulum;Endosome;Membrane;Nucleus', u'Cdc37l1': 'Cytoplasm', u'Riok2': '', u'Igfbp2': 'Secreted', u'Igfbp1': 'Secreted', u'Igfbp6': 'Secreted', u'Casp8': 'Cytoplasm', u'Igfbp5': 'Secreted', u'Ppib': 'Endoplasmic reticulum', u'Ppia': 'Cytoplasm;Secreted', u'Hpd': 'Cytoplasm;Endoplasmic reticulum;Golgi apparatus;Membrane', u'Igf1r': 'Cell membrane;Membrane', u'NSF': '', u'Nup153': 'Membrane;Nuclear pore complex;Nucleus', u'Grb2': 'Cytoplasm;Endosome;Golgi apparatus;Nucleus', u'Chdh': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'pam': '', u'B3galt5': 'Golgi apparatus', u'Acaca': 'Cytoplasm', u'Taf5': '', u'slc8a1': '', u'Itln1': '', u'Uso1': 'Cytoplasm;Golgi apparatus;Membrane', u'Cep112': '', u'Blnk': 'Cell membrane;Cytoplasm;Membrane', u'RGD1566137': '', u'RGD1566136': '', u'Cgn': '', u'RGD1565368': '', u'ENSRNOG00000042516': '', u'Eif3s6ip': '', u'Cox6a1': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'LOC689899': '', u'Cox6a2': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'Ncapg2': '', u'Acsl3': 'Endoplasmic reticulum;Membrane;Microsome;Mitochondrion;Mitochondrion outer membrane;Peroxisome', u'Hnrnpr': 'Virion', u'Acsl1': 'Endoplasmic reticulum;Membrane;Microsome;Mitochondrion;Mitochondrion outer membrane;Peroxisome', u'Acsl6': 'Endoplasmic reticulum;Membrane;Microsome;Mitochondrion;Mitochondrion outer membrane;Peroxisome', u'Acsl4': '', u'Acsl5': 'Endoplasmic reticulum;Membrane;Mitochondrion;Mitochondrion outer membrane', u'Erp44': '', u'Hnrnpk': 'Cell junction;Cell projection;Cytoplasm;Nucleus;Spliceosome', u'Hnrnpl': 'Cytoplasm;Nucleus', u'Hnrnpm': 'Nucleus;Spliceosome', u'Hnrnpf': 'Nucleus;Spliceosome', u'Stk24': 'Cytoplasm;Membrane;Nucleus', u'Tnnc2': '', u'LYN': '', u'Engase': '', u'Sh3tc2': '', u'Mdh2': 'Mitochondrion', u'Mdh1': 'Cytoplasm', u'Pin1': '', u'Dicer1': '', u'Dnajc5': 'Cell membrane;Membrane', u'Dnajc6': '', u'cand1': '', u'Dnajc3': 'Endoplasmic reticulum', u'Agrn': 'Cell junction;Cell membrane;Membrane;Synapse', u'Bin2a': '', u'Il4i1': '', u'Kif3a': 'Microtubule', u'Ngrn': 'Membrane;Mitochondrion;Nucleus;Secreted', u'Racgap1': '', u'Ppp2r2d': 'Cytoplasm', u'HLA-A': '', u'Ppp2r2a': '', u'Ppp2r2c': '', u'Ppp2r2b': 'Cytoplasm;Cytoskeleton;Membrane;Mitochondrion;Mitochondrion outer membrane', u'CLTC': '', u'Mars': '', u'ENSRNOG00000048597': '', u'Psmb4': 'Cytoplasm;Nucleus;Proteasome', u'Met': 'Membrane', u'Pgam2': '', u'Rptor': '', u'Pabpc1': 'Cytoplasm;Nucleus;Spliceosome', u'Gpihbp1': '', u'pdia4': '', u'Fndc1': 'Secreted', u'Gtf2b': 'Chromosome;Nucleus', u'Hist1h2ao': '', u'e3_vaccw': '', u'Plp1': 'Cell membrane;Membrane', u'Nat10': 'Nucleus', u'Rab4a': 'Cytoplasm;Endosome;Membrane', u'LOC685069': '', u'Cnmpd1': '', u'Me1': 'Cytoplasm', u'Me3': '', u'Me2': '', u'BAIAP2': '', u'Ndufa6': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'Sec13': 'Cytoplasmic vesicle;Endoplasmic reticulum;Lysosome;Membrane;Nuclear pore complex;Nucleus', u'Ndufa8': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'Adm': 'Secreted', u'ENSRNOG00000033834': '', u'ENSRNOG00000048625': '', u'Ccdc71': '', u'Slc32a1': 'Cytoplasmic vesicle;Membrane', u'Gapdhs': 'Cytoplasm', u'Gpbp1': '', u'Spata5l1': '', u'Gapdh-ps2': '', u'B3galnt1': 'Golgi apparatus;Membrane', u'Tspan10': 'Membrane', u'Rasl2-9': 'Nucleus', u'Tnf': 'Cell membrane;Membrane;Secreted', u'Insrr': 'Membrane', u'ENSRNOG00000047202': '', u'Tbc1d15': '', u'ENSRNOG00000047114': '', u'Eno2': 'Cell membrane;Cytoplasm;Membrane', u'Rabep2': 'Cytoplasm;Endosome', u'Nptxr': 'Membrane', u'Rabep1': 'Cytoplasm;Cytoplasmic vesicle;Endosome', u'Tmem2': '', u'Rnaset2': '', u'atxn3': '', u'Rps26': 'Cytoplasm;Endoplasmic reticulum', u'Rps27': '', u'Rragd': '', u'Rps25': '', u'Srsf1': '', u'Rps23': 'Cytoplasm;Endoplasmic reticulum', u'Wnt2': 'Extracellular matrix', u'Srsf9': 'Nucleus', u'Rps28': 'Cytoplasm;Endoplasmic reticulum', u'Rps29': 'Cytoplasm;Endoplasmic reticulum', u'Dhtkd1': 'Mitochondrion', u'RGD1560402': '', u'Nfkbia': 'Cytoplasm;Nucleus', u'Arhgdib': '', u'Arhgdia': 'Cytoplasm', u'Atp6v1b1': '', u'Atp6v1b2': 'Membrane', u'Actr1b': '', u'Degs1': 'Endoplasmic reticulum;Membrane;Mitochondrion', u'Degs2': 'Endoplasmic reticulum;Membrane', u'Sugt1': 'Cytoplasm;Nucleus', u'Vldlr': 'Coated pit;Membrane;VLDL', u'Prkca': 'Cell membrane;Cytoplasm;Membrane;Mitochondrion;Nucleus', u'Prkcb': 'Cytoplasm;Membrane;Nucleus', u'Prkar1a': 'Cell membrane;Membrane', u'Prkcg': 'Cell junction;Cell membrane;Cell projection;Cytoplasm;Membrane;Synapse;Synaptosome', u'Prkcq': 'Cell membrane;Cytoplasm;Membrane', u'PACS1': '', u'Neurl1B': '', u'rph3al': '', u'Atp1b4': 'Membrane;Nucleus', u'Sep15': '', u'Atp1b2': 'Cell membrane;Membrane', u'Atp1b3': 'Cell membrane;Membrane', u'Atp1b1': 'Cell membrane;Membrane', u'Jak2': 'Cytoplasm;Membrane;Nucleus', u'Atg7': 'Cytoplasm', u'Arfgap1': 'Cytoplasm;Golgi apparatus', u'US9': '', u'Lrrc28': '', u'Uox': 'Peroxisome', u'Ccnj': '', u'ENSRNOG00000030426': '', u'Snta1': '', u'Iffo2': 'Intermediate filament', u'Cyb561': 'Membrane', u'Hnrnpa1': 'Cytoplasm;Nucleus;Spliceosome', u'Rin1': 'Cytoplasm;Cytoskeleton;Membrane', u'Mrpl4': '', u'GLP1R': '', u'map1lc3a': '', u'tg': '', u'map1lc3b': '', u'Ptprn2': 'Cell junction;Cytoplasmic vesicle;Membrane;Synapse', u'Ptdss1': 'Endoplasmic reticulum;Membrane', u'Asah1': 'Lysosome', u'Rasgrp3': '', u'ATP6': '', u'Cdc5l': 'Cytoplasm;Nucleus;Spliceosome', u'Kpna6': '', u'Rabac1': 'Cell junction;Cell membrane;Cytoplasm;Cytoplasmic vesicle;Golgi apparatus;Membrane;Synapse', u'Tnip3': '', u'Foxa2': 'Cytoplasm;Nucleus', u'ENSRNOG00000047799': '', u'Kpna1': 'Cytoplasm;Nucleus', u'Wbscr16': '', u'Gp1ba': 'Membrane', u'Gp1bb': 'Membrane', u'Mkks': '', u'Ap2b1': 'Cell membrane;Coated pit;Membrane', u'Gad2': 'Cell junction;Cell membrane;Cytoplasm;Cytoplasmic vesicle;Golgi apparatus;Membrane;Synapse', u'Bcl2l1': 'Cell junction;Cytoplasm;Cytoplasmic vesicle;Cytoskeleton;Membrane;Mitochondrion;Mitochondrion inner membrane;Mitochondrion outer membrane;Nucleus;Synapse', u'Dnajb9': 'Endoplasmic reticulum', u'Dnajb3': '', u'Dnajb2': '', u'Dnajb6': 'Cytoplasm;Nucleus', u'Apoe': 'Chylomicron;HDL;Secreted;VLDL', u'Dynll1': 'Cytoplasm;Cytoskeleton;Dynein;Microtubule;Mitochondrion;Nucleus', u'Prkab1': '', u'Srrm1': '', u'tuba1a': '', u'Camsap1': 'Cytoplasm;Cytoskeleton;Microtubule', u'Ptcd3': '', u'HCK': '', u'Psmc2': 'Cytoplasm;Proteasome', u'cyth2': '', u'Psmc4': 'Cytoplasm;Nucleus;Proteasome', u'Psmc5': 'Cytoplasm;Nucleus;Proteasome', u'Psmc6': '', u'LOC690675': '', u'nedd4': '', u'Hsp90b1': 'Endoplasmic reticulum', u'Atp6v0a4': 'Membrane', u'Atp6v0a1': 'Cytoplasmic vesicle;Membrane', u'calr': '', u'Canx': 'Endoplasmic reticulum;Membrane', u'N4bp2': '', u'Arl9': '', u'Atp5a1': '', u'Mrps30': '', u'Clip1': 'Cell projection;Cytoplasm;Cytoplasmic vesicle;Cytoskeleton;Membrane;Microtubule', u'Vti1a': 'Cell junction;Cytoplasmic vesicle;Golgi apparatus;Membrane;Synapse', u'Snd1': 'Cytoplasm;Nucleus', u'Vti1b': 'Endosome;Lysosome;Membrane', u'Aco1': 'Cytoplasm', u'Aco2': 'Mitochondrion', u'Sel1l': 'Endoplasmic reticulum;Membrane', u'ENSRNOG00000048058': '', u'Plaur': 'Cell membrane;Membrane;Secreted', u'Ago2': 'Cytoplasm', u'Crispld1': '', u'Supt16h': '', u'C1qbp': 'Cell membrane;Cytoplasm;Membrane;Mitochondrion;Nucleus;Secreted', u'Si': 'Cell membrane;Membrane', u'Ggt1': 'Cell membrane;Membrane', u'LTF': '', u'Casp9': '', u'Serpine1': 'Secreted', u'Ahcyl1': 'Cell membrane;Cytoplasm;Endoplasmic reticulum;Membrane;Microsome', u'Dhfr': 'Cytoplasm;Mitochondrion', u'ntrk1': '', u'Clcf1': '', u'Spata5': '', u'Ak8': 'Cytoplasm', u'Igfbp4': 'Secreted', u'Rab3ip': 'Cell projection;Cytoplasm;Cytoskeleton;Nucleus', u'Ak2': 'Mitochondrion', u'Ak1': 'Cytoplasm', u'ENSRNOG00000047279': '', u'Ddx23': '', u'Ddx24': '', u'Sparc': 'Basement membrane;Extracellular matrix;Secreted', u'Ssr4': 'Endoplasmic reticulum;Membrane', u'mapt': '', u'CXCR4': '', u'Dennd1c': '', u'Dennd1b': '', u'actb': '', u'cdc42': '', u'Rps11': '', u'Rps17': '', u'Rps16': '', u'Rps15': '', u'Rps14': '', u'Gla': '', u'Pfkp': 'Cytoplasm', u'Slc12a4': 'Membrane', u'Pfkm': 'Cytoplasm', u'Pfkl': 'Cytoplasm', u'Jagn1': 'Endoplasmic reticulum;Membrane', u'Rln3': 'Secreted', u'Cwc25': '', u'B4galnt1': 'Golgi apparatus;Membrane', u'calm1': '', u'Acan': 'Extracellular matrix;Secreted', u'Birc5': 'Centromere;Chromosome;Cytoplasm;Cytoskeleton;Kinetochore;Microtubule;Nucleus', u'Mylk2': 'Cytoplasm', u'Birc2': '', u'Gnptab': 'Membrane', u'Unc13d': 'Cytoplasm;Endosome;Lysosome;Membrane', u'Atp6v1c2': '', u'Atp6v1c1': '', u'Rap1gap2': '', u'Akr1b10': '', u'Lgmn': 'Lysosome', u'Tuba4a': 'Cytoplasm;Cytoskeleton;Microtubule', u'Prcp': '', u'Hist2h2ab': 'Chromosome;Nucleosome core', u'Anxa2': 'Basement membrane;Extracellular matrix;Secreted', u'Lrrc58': '', u'Smarca4': 'Nucleus', u'Atf3': 'Nucleus', u'c6-cer': '', u'Pklr': '', u'Serpina1': 'Secreted', u'gnao1': '', u'Snupn': 'Cytoplasm;Nucleus', u'Bax': 'Cytoplasm;Membrane;Mitochondrion;Mitochondrion outer membrane', u'Bad': 'Cytoplasm;Membrane;Mitochondrion;Mitochondrion outer membrane', u'Faslg': 'Cell membrane;Cytoplasmic vesicle;Lysosome;Membrane;Nucleus;Secreted', u'Gdi2': 'Cytoplasm;Membrane', u'Prc1': '', u'Gdi1': 'Cytoplasm;Golgi apparatus', u'Eny2': 'Nuclear pore complex', u'Mut': '', u'Eng': 'Membrane', u'ENSRNOG00000028543': '', u'Eno1': 'Cell membrane;Cytoplasm;Membrane', u'Tpi1': '', u'Eno3': 'Cytoplasm', u'nptxr': '', u'Pa2g4': 'Cytoplasm;Nucleus', u'Ap2s1': 'Cell membrane;Coated pit;Membrane', u'Miro': '', u'Becn1': 'Cytoplasm;Cytoplasmic vesicle;Endoplasmic reticulum;Endosome;Golgi apparatus;Membrane;Mitochondrion;Nucleus', u'Irs2': '', u'Aurkb': 'Centromere;Chromosome;Cytoplasm;Cytoskeleton;Nucleus', u'Abca12': 'Membrane', u'Tmsb4x': 'Cytoplasm;Cytoskeleton', u'Arhgef3': '', u'LOC680700': '', u'nptx2': '', u'Eif2s3x': '', u'Tas1r1': 'Cell membrane;Membrane', u'Tas1r3': 'Cell membrane;Membrane', u'Cct6a': 'Cytoplasm', u'Cct6b': 'Cytoplasm', u'Fxyd7': 'Membrane', u'Fxyd2': 'Membrane', u'Fermt3': '', u'ENSRNOG00000049269': '', u'Pkm': 'Cytoplasm;Nucleus', u'Derl1': 'Endoplasmic reticulum', u'Derl2': 'Endoplasmic reticulum', u'Derl3': 'Endoplasmic reticulum', u'ENSRNOG00000043026': '', u'RGD1562690': '', u'Lrp2': 'Cell membrane;Cell projection;Coated pit;Endosome;Membrane', u'gnb1': '', u'Pink1': '', u'ENSRNOG00000023237': '', u'Kif4a': '', u'Taf13': '', u'Taf12': '', u'Taf10': 'Nucleus', u'Map6': 'Cell projection;Cytoplasm;Cytoplasmic vesicle;Cytoskeleton;Golgi apparatus;Membrane;Microtubule', u'Pik3r1': '', u'Prkacb': 'Cell membrane;Cytoplasm;Membrane;Nucleus', u'Tars2': 'Mitochondrion', u'Rab6a': 'Golgi apparatus;Membrane', u'Ipo7': 'Nucleus', u'Dis3l2': 'Cytoplasm', u'stx1a': '', u'Lclat1': 'Membrane', u'Cyb561d1': 'Membrane', u'Gabrb2': 'Cell junction;Cell membrane;Cytoplasmic vesicle;Membrane;Postsynaptic cell membrane;Synapse', u'Ddx3x': '', u'Adamtsl5': '', u'Rplp1': '', u'Rplp0': 'Cytoplasm;Nucleus', u'Kit': 'Cell membrane', u'Srgn': 'Golgi apparatus;Secreted', u'Acrbp': 'Cytoplasmic vesicle;Secreted', u'Mthfd1l': '', u'Dhx15': '', u'ENSRNOG00000045756': '', u'Mthfd2': '', u'Mthfd1': 'Cytoplasm', u'Usp10': 'Cytoplasm;Endosome;Nucleus', u'Prkaca': 'Cell membrane;Cell projection;Cilium;Cytoplasm;Cytoplasmic vesicle;Flagellum;Membrane;Mitochondrion;Nucleus', u'Stip1': 'Cytoplasm;Nucleus', u'Cdc42': 'Cell membrane;Cell projection;Cytoplasm;Cytoskeleton;Membrane', u'Slc39a13': 'Golgi apparatus;Membrane', u'Tm6sf1': 'Membrane', u'Nsun3': '', u'Adrbk1': '', u'Cwc15': 'Nucleus;Spliceosome', u'CAV1': '', u'Ppfia1': '', u'Rph3a': 'Cell junction;Membrane;Synapse', u'Slc30a9': 'Membrane', u'RGD1561609': '', u'Rabl6': '', u'Slc30a7': 'Golgi apparatus;Membrane', u'Slc30a6': 'Membrane', u'Tnnt2': '', u'Psap': 'Lysosome;Secreted', u'Frmd4a': '', u'Mpst': 'Cell junction;Cytoplasm;Mitochondrion;Synapse;Synaptosome', u'Lrrc40': '', u'Lrrc47': '', u'Lpin1': '', u'Eef1g': '', u'Eef1d': 'Nucleus', u'parp1': '', u'LOC301861': '', u'arc': '', u'Guk1': '', u'RGD1561333': '', u'Mon1a': '', u'Furin': 'Cell membrane;Endosome;Golgi apparatus;Membrane;Secreted', u'34875531_rat_prot': '', u'Dennd2d': '', u'Bub1b': '', u'epn1': '', u'Dvl1': 'Cell membrane;Cytoplasm;Cytoplasmic vesicle;Membrane', u'Plg': 'Secreted', u'Pln': 'Endoplasmic reticulum;Membrane;Mitochondrion;Sarcoplasmic reticulum', u'Gcfc2': '', u'Mgam': '', u'Chp1': 'Cell membrane;Cytoplasm;Cytoskeleton;Endoplasmic reticulum;Membrane;Nucleus', u'Gal3st1': 'Membrane', u'Cdca7l': 'Cytoplasm;Nucleus', u'Gal3st4': 'Membrane', u'Myl9': '', u'Sac3d1': '', u'Tomm40': 'Membrane;Mitochondrion;Mitochondrion outer membrane', u'Itgb7': 'Membrane', u'Calcrl': 'Cell membrane;Membrane', u'Gorasp2': 'Golgi apparatus;Membrane', u'rCG_56798': '', u'Tlr9': '', u'Nos1ap': '', u'Nckipsd': '', u'Dnajc27': 'Nucleus', u'Whsc1': '', u'Kif5c': 'Cell projection;Cytoplasm;Cytoskeleton;Microtubule', u'Mpi': 'Cytoplasm', u'Atp5hl1': '', u'Hspa2': 'Cytoplasm;Cytoskeleton', u'Hspa1': '', u'Agt': 'Secreted', u'Slc44a4': 'Cell membrane;Membrane', u'Slc44a5': 'Membrane', u'Hspa5': 'Endoplasmic reticulum', u'Hspa4': 'Cytoplasm', u'Pik3cb': 'Cytoplasm;Nucleus', u'Hspa8': 'Cell membrane;Cytoplasm;Membrane;Nucleus;Spliceosome', u'Dnajb12': 'Membrane', u'Dnajb11': 'Endoplasmic reticulum', u'Sept12': 'Cell projection;Cilium;Cytoplasm;Cytoskeleton;Flagellum', u'A2m': 'Secreted', u'Mtr': 'Cytoplasm', u'Flt4': 'Cell membrane;Cytoplasm;Membrane;Nucleus', u'Atxn7l3': 'Nucleus', u'Flt3': 'Membrane', u'Cox5a': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'Tubal3': 'Cytoplasm;Cytoskeleton', u'Cuta': '', u'Got2': 'Cell membrane;Membrane;Mitochondrion', u'Ptpn2': 'Cell membrane;Cytoplasm;Endoplasmic reticulum;Membrane;Nucleus', u'Rab13': 'Cell junction;Cell membrane;Cell projection;Cytoplasmic vesicle;Endosome;Golgi apparatus;Membrane;Tight junction', u'Pik3c3': 'Cytoplasmic vesicle;Endosome', u'Pgk2': '', u'PPIB': '', u'Pgk1': 'Cytoplasm', u'Pls1': '', u'Vma21': '', u'Il6': 'Secreted', u'Il2': 'Secreted', u'ENSRNOG00000016319': '', u'Nploc4': 'Cytoplasm;Endoplasmic reticulum;Nucleus', u'Copa': 'Cytoplasm', u'Klhl17': 'Cell junction;Cell membrane;Membrane;Postsynaptic cell membrane;Synapse', u'Tmem231': 'Cell membrane;Cell projection;Cilium;Membrane', u'Eif2b5': '', u'Eif2b3': '', u'Actr2': 'Cell projection;Cytoplasm;Cytoskeleton', u'Ddx5': '', u'EBI-6372956': '', u'ENSRNOG00000033686': '', u'Ndufb9': '', u'Ndufb7': '', u'Sptan1': 'Cytoplasm;Cytoskeleton', u'Ndufb5': 'Membrane', u'ENSRNOG00000000574': '', u'Smarcal1': 'Nucleus', u'Serping1': 'Secreted', u'Fst': 'Secreted', u'Galnt12': 'Golgi apparatus', u'vdac1': '', u'Raf1': 'Cell membrane;Cytoplasm;Membrane;Mitochondrion;Nucleus', u'ENSRNOG00000048142': '', u'Sil1': 'Endoplasmic reticulum', u'amph': '', u'ENSRNOG00000047670': '', u'Kpnb1': 'Cytoplasm;Nucleus', u'Tapbp': 'Membrane', u'akt1': '', u'ENSRNOG00000047070': '', u'Manf': 'Endoplasmic reticulum;Sarcoplasmic reticulum;Secreted', u'ENSRNOG00000028339': '', u'Opa1': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'EBNA-LP': '', u'Myo5b': 'Cytoplasm', u'Xpnpep2': 'Cell membrane;Membrane', u'Rab3il1': '', u'Sgk1': 'Cell membrane;Cytoplasm;Endoplasmic reticulum;Membrane;Mitochondrion;Nucleus', u'Ctsa': '', u'Ctsb': 'Lysosome;Secreted', u'Vav2': '', u'Ctsd': 'Lysosome;Secreted', u'Ctse': 'Endosome', u'Tcirg1': 'Membrane', u'Ero1lb': '', u'Ctsz': 'Secreted', u'Akr1b8': '', u'Hgsnat': '', u'SEPT14': '', u'Ctss': 'Lysosome', u'Gpr174': 'Membrane', u'Gns': 'Lysosome', u'Ganab': '', u'Ccna2': '', u'siah1': '', u'Myrip': 'Cytoplasm;Cytoplasmic vesicle', u'Rph3al': 'Cytoplasm;Cytoplasmic vesicle;Membrane', u'Cenpe': '', u'Mapk3': 'Cytoplasm;Membrane;Nucleus', u'Mapk1': 'Cytoplasm;Cytoskeleton;Membrane;Nucleus', u'Gne': 'Cytoplasm', u'Cldn1': 'Cell junction;Cell membrane;Membrane;Tight junction', u'Eprs': '', u'p97582_rat': '', u'Ckap5': '', u'Nudt8': '', u'Ptges3': 'Cytoplasm', u'Smpd4': 'Membrane', u'katna1': '', u'Smpd1': 'Membrane', u'Pcdhgb5': 'Membrane', u'Smpd2': 'Membrane', u'KDELR1': '', u'fxyd7': '', u'Gnaq': 'Membrane;Nucleus', u'Idh2': 'Mitochondrion', u'Idh1': 'Cytoplasm;Peroxisome', u'Gnaz': 'Membrane', u'Impa1': 'Cytoplasm', u'Napa': 'Cell membrane;Membrane', u'arfgap1': '', u'Gnal': '', u'8-aha-omethyladenosine': '', u'Os9': 'Endoplasmic reticulum', u'Gna15': '', u'Gna12': 'Cell membrane;Cytoplasm;Membrane', u'Gna13': 'Cytoplasm;Membrane;Nucleus', u'G3bp1': '', u'Uggt1': 'Endoplasmic reticulum', u'Fam171a2': 'Membrane', u'Tjp2': '', u'Hyou1': 'Endoplasmic reticulum', u'Plod3': 'Endoplasmic reticulum;Membrane', u'Plod1': 'Endoplasmic reticulum;Membrane', u'Gm8420': '', u'Asns': '', u'Lao1': '', u'Myc': 'Nucleus', u'Cox4i2': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'Cox4i1': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'mrs6': '', u'Klc1': 'Cell projection;Cytoplasm;Cytoplasmic vesicle;Cytoskeleton;Microtubule', u'Klc2': '', u'Klc3': 'Cytoplasm;Cytoskeleton;Microtubule', u'Nucb1': 'Cytoplasm;Golgi apparatus;Membrane;Secreted', u'Snap25': 'Cell junction;Cell membrane;Cytoplasm;Membrane;Synapse;Synaptosome', u'Entpd4': 'Membrane', u'rad23b': '', u'Snap23': 'Cell junction;Cell membrane;Cytoplasmic vesicle;Membrane;Synapse;Synaptosome', u'bves': '', u'ENSRNOG00000030871': '', u'Pi4k2a': 'Cell junction;Cell membrane;Cell projection;Cytoplasmic vesicle;Endosome;Golgi apparatus;Membrane;Mitochondrion;Synapse;Synaptosome', u'Snap29': 'Cell membrane;Cell projection;Cilium;Cytoplasm;Cytoplasmic vesicle;Golgi apparatus;Membrane', u'Map1lc3a': 'Cytoplasm;Cytoplasmic vesicle;Cytoskeleton;Membrane;Microtubule', u'Plxna4a': '', u'Efnb2': 'Membrane', u'Efnb3': 'Membrane', u'Efnb1': 'Membrane', u'Cyct': 'Mitochondrion', u'Prdx5': 'Cytoplasm;Mitochondrion;Peroxisome', u'Prdx2': 'Cytoplasm', u'Prdx1': 'Cytoplasm', u'Tgoln2': '', u'Leo1': 'Nucleus', u'TRAK1': '', u'Pomgnt1': 'Golgi apparatus;Membrane', u'Dnajc30': 'Membrane', u'P4ha2': '', u'Fh': 'Cytoplasm;Mitochondrion', u'P4ha1': 'Endoplasmic reticulum', u'Napsa': '', u'Hnrpd': '', u'Cyc1': '', u'PDE4DIP': '', u'Hsp90aa1': 'Cell membrane;Cytoplasm;Membrane;Nucleus', u'Egln3': 'Cytoplasm;Nucleus', u'Ero1l': '', u'Prkaa2': 'Cytoplasm;Nucleus', u'Prkaa1': 'Cytoplasm;Nucleus', u'Fxr1': 'Cytoplasm', u'Th': '', u'Tf': 'Secreted', u'Tg': 'Secreted', u'Taf7': '', u'Map4k3': '', u'Max': 'Cell projection;Nucleus', u'Ufd1l': '', u'Map2k2': 'Cytoplasm;Membrane', u'Map2k1': 'Cytoplasm;Cytoskeleton;Membrane;Nucleus', u'Map2k5': 'Cytoplasm;Membrane', u'Edn1': 'Secreted', u'EBI-2257696': '', u'csnk2a1': '', u'Nagpa': 'Membrane', u'Mcf2': '', u'Eif5': '', u'Glud1': 'Mitochondrion', u'RGD1563962': '', u'Sf3b2': '', u'ENSRNOG00000032585': '', u'Eif2c1': '', u'Eif2c2': '', u'Dpp8': '', u'Dpp9': '', u'Rapgef1': '', u'Actr1a': 'Cytoplasm;Cytoskeleton', u'Dpp4': 'Cell junction;Cell membrane;Cell projection;Membrane;Secreted', u'Dpp7': 'Cytoplasmic vesicle;Lysosome;Secreted', u'Gip': 'Secreted', u'Islr': '', u'Galns': 'Lysosome', u'Kpna7': '', u'Fbxo22': '', u'Kpna5': 'Cytoplasm', u'Kpna4': '', u'Kpna3': '', u'Kpna2': '', u'FKBP2': '', u'Atp6ap1l': 'Membrane', u'rac1': '', u'Slc39a8': 'Membrane', u'vegfa': '', u'Clca2': 'Membrane', u'Ogdhl': '', u'Slc39a4': 'Cell membrane;Endosome;Membrane', u'Slc39a6': 'Cell membrane;Membrane', u'Clca5': '', u'Sumf2': '', u'Sumf1': '', u'Endog': '', u'ENSRNOG00000042227': '', u'Dhx38': '', u'RGD1559493': '', u'Gngt1': 'Cell membrane', u'Trpm7': 'Membrane', u'Nipal3': 'Membrane', u'Ccnb2': '', u'Ccnb1': 'Cytoplasm;Cytoskeleton;Nucleus', u'Hexa': 'Lysosome', u'RT1-Ba': 'Membrane;MHC II', u'RT1-Bb': 'Membrane;MHC II', u'Hexb': 'Lysosome', u'Flot1': 'Cell membrane;Endosome;Membrane', u'Gabrb1': 'Cell junction;Cell membrane;Membrane;Postsynaptic cell membrane;Synapse', u'Wipf1': 'Cell projection;Cytoplasm;Cytoplasmic vesicle;Cytoskeleton', u'Gabrb3': 'Cell junction;Cell membrane;Cytoplasmic vesicle;Membrane;Postsynaptic cell membrane;Synapse', u'Lama3': '', u'Iqgap1': '', u'Pfn1': 'Cytoplasm;Cytoskeleton', u'8-aha-camp': '', u'Dennd2c': '', u'clic4': '', u'Dennd2a': '', u'cnn1': '', u'hspa5': '', u'hspa9': '', u'Arrb1': 'Cell membrane;Cell projection;Coated pit;Cytoplasm;Cytoplasmic vesicle;Membrane;Nucleus', u'Uhmk1': 'Cytoplasm;Nucleus', u'Slc7a11': 'Membrane', u'GDI1': '', u'Ldhb': 'Cytoplasm', u'Ldhc': 'Cytoplasm', u'Gm2a': '', u'Rps6kb1': 'Cell junction;Cytoplasm;Membrane;Mitochondrion;Mitochondrion outer membrane;Synapse;Synaptosome', u'5-nh2_caproyl-isatin': '', u'RGD1561481': '', u'Rbms2': '', u'Pip5k1a': 'Cell membrane;Cell projection;Cytoplasm;Membrane;Nucleus', u'Ube4b': '', u'Ntf3': 'Secreted', u'ENSRNOG00000029533': '', u'Pisd': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'Rxfp3': 'Membrane', u'Arcn1': 'Cytoplasm;Cytoplasmic vesicle;Golgi apparatus;Membrane', u'Cul4a': '', u'Cul4b': '', u'Tubb2a': 'Cytoplasm;Cytoskeleton;Microtubule', u'Tubb2b': 'Cytoplasm;Cytoskeleton;Microtubule', u'Tuba3a': 'Cytoplasm;Cytoskeleton;Microtubule', u'Tjp1': 'Cell junction;Cell membrane;Gap junction;Membrane;Tight junction', u'Hap1': 'Cell junction;Cell projection;Cytoplasm;Cytoplasmic vesicle;Cytoskeleton;Endoplasmic reticulum;Lysosome;Mitochondrion;Nucleus;Synapse', u'Solh': '', u'Pnp': 'Cytoplasm;Cytoskeleton', u'Nadsyn1': '', u'brdt': '', u'Idua': '', u'Eif3h': 'Cytoplasm', u'Eif3i': 'Cytoplasm', u'Eif3j': 'Cytoplasm', u'Eif3k': 'Cytoplasm', u'Zmynd8': '', u'Eif3m': 'Cytoplasm', u'Eif3a': 'Cytoplasm', u'Eif3b': 'Cytoplasm', u'Eif3c': 'Cytoplasm', u'Eif3d': 'Cytoplasm', u'Eif3e': 'Cytoplasm;Nucleus', u'Eif3g': 'Cytoplasm;Nucleus', u'Syvn1': 'Membrane', u'Cfl2': '', u'Cfl1': 'Cell membrane;Cell projection;Cytoplasm;Cytoskeleton;Membrane;Nucleus', u'S100a6': 'Cell membrane;Cytoplasm;Membrane;Nucleus', u'F2': '', u'F3': 'Membrane', u'F5': '', u'Timp1': 'Secreted', u'Efna5': 'Cell membrane;Membrane', u'Efna4': 'Membrane', u'Efna3': 'Membrane', u'Efna2': 'Membrane', u'Efna1': 'Cell membrane;Membrane;Secreted', u'sptbn1': '', u'Pla2g6': 'Cytoplasm;Membrane', u'Rpl10a': '', u'Dynlt3': '', u'Pi4kb': 'Endoplasmic reticulum;Golgi apparatus;Membrane;Mitochondrion;Mitochondrion outer membrane', u'Limk1': 'Cell projection;Cytoplasm;Nucleus', u'Glb1': '', u'RSA-14-44': '', u'Rab3d': 'Cell membrane;Membrane', u'Rab3c': 'Cell membrane;Membrane', u'Rab3b': 'Cell membrane;Golgi apparatus;Membrane', u'Rab3a': 'Cell membrane;Membrane', u'RGD1309821': '', u'bin1': '', u'ENSRNOG00000028664': '', u'Gapdh': 'Cytoplasm;Cytoskeleton;Nucleus', u'Lgals3bp': 'Extracellular matrix;Secreted', u'Rab11b': 'Cell junction;Cytoplasmic vesicle;Endosome;Membrane;Synapse', u'Sct': 'Secreted', u'Pla2g4b': 'Cytoplasm', u'Stx8': 'Membrane', u'Ugcg': 'Golgi apparatus;Membrane', u'ENSRNOG00000004593': '', u'Rpl27a': '', u'Stx2': 'Membrane', u'Stx5': 'Golgi apparatus;Membrane', u'Stx4': 'Cell membrane;Membrane', u'Stx7': 'Endosome;Membrane', u'Stx6': 'Golgi apparatus;Membrane', u'Spock3': '', u'Veph1': 'Cell membrane;Membrane', u'Mthfd2l': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'tanc1': '', u'Asap1': 'Cytoplasm;Membrane', u'Pmpcb': 'Mitochondrion', u'Setmar': 'Chromosome;Nucleus', u'Sst': 'Secreted', u'atp1a3': '', u'Gp2': 'Cell membrane;Membrane;Secreted', u'atp1a1': '', u'Lamb1': '', u'Lamb2': 'Basement membrane;Extracellular matrix;Secreted', u'Cd320': 'Cell membrane;Membrane', u'Ints10': '', u'NCKIPSD': '', u'Atox1': '', u'Mmp9': 'Extracellular matrix;Secreted', u'Cat': 'Peroxisome', u'ENSRNOG00000023115': '', u'Sigmar1': 'Cell junction;Cell membrane;Cell projection;Cytoplasmic vesicle;Endoplasmic reticulum;Lipid droplet;Membrane;Nucleus;Postsynaptic cell membrane;Synapse', u'Cad': '', u'Sema3a': 'Secreted', u'Psapl1': '', u'Clec3b': '', u'GNG2': '', u'mapk7': '', u'Gart': '', u'Rab1': '', u'Gars': 'Cell projection;Cytoplasm;Secreted', u'Atp6v1d': '', u'Timmdc1': 'Membrane;Mitochondrion', u'Atp6v1f': '', u'Atp6v1a': '', u'RGD1565317': '', u'Atp6v1h': '', u'Tia1': '', u'Ptges3l': '', u'Srxn1': '', u'Chgb': 'Secreted', u'Arf4': 'Golgi apparatus;Membrane', u'Ubac2': 'Membrane', u'Rabif': '', u'Ace2': 'Cell membrane;Cytoplasm;Membrane;Secreted', u'Cdkal1': '', u'Tmed7': 'Cytoplasmic vesicle;Endoplasmic reticulum;Golgi apparatus;Membrane', u'Tmed3': 'Cytoplasmic vesicle;Endoplasmic reticulum;Golgi apparatus;Membrane', u'Tmed2': 'Cytoplasmic vesicle;Endoplasmic reticulum;Golgi apparatus;Membrane', u'Ghr': 'Cell membrane;Membrane;Secreted', u'2_aha_camp': '', u'Tmed9': 'Endoplasmic reticulum;Golgi apparatus;Membrane', u'Pdpk1': 'Cell junction;Cell membrane;Cytoplasm;Membrane;Nucleus', u'Rab11a': 'Cell membrane;Cytoplasmic vesicle;Endosome;Membrane', u'hyou1': '', u'Ssrp1': 'Chromosome;Nucleus', u'Tkt': '', u'ENSRNOG00000042129': '', u'Adsl': '', u'Ehhadh': 'Peroxisome', u'Abcg8': 'Cell membrane;Membrane', u'Proc': 'Endoplasmic reticulum;Golgi apparatus;Secreted', u'Mep1a': 'Membrane', u'Mep1b': 'Cell membrane;Membrane;Secreted', u'Sucla2': 'Mitochondrion', u'Crisp2': '', u'Sar1b': 'Endoplasmic reticulum;Golgi apparatus;Membrane', u'ENSRNOG00000029446': '', u'Vgf': 'Cytoplasmic vesicle;Secreted', u'Capza2': '', u'Deaf1': 'Nucleus;Secreted', u'Col6a4': '', u'Rab35': 'Cell membrane;Coated pit;Cytoplasmic vesicle;Endosome;Membrane', u'Mylk': '', u'LAMP3': '', u'zbtb16': '', u'Cln6': 'Membrane', u'Cln5': '', u'Cln3': 'Membrane', u'Stx3': 'Membrane', u'Cln8': 'Endoplasmic reticulum;Membrane', u'Erp27': '', u'Erp29': 'Endoplasmic reticulum', u'Atg12': 'Cytoplasm;Membrane', u'ENSRNOG00000027267': '', u'CYTB': 'Membrane;Mitochondrion', u'Eif1ax': '', u'Uba52': 'Cytoplasm;Nucleus', u'Col15a1': '', u'Rpl5': 'Cytoplasm;Nucleus', u'Rraga': 'Cytoplasm;Lysosome;Nucleus', u'Ccdc132': '', u'Ppa2': '', u'Ppa1': '', u'Myoz2': '', u'Slc1a7': 'Membrane', u'Col18a1': '', u'Olr853': 'Cell membrane', u'park7': '', u'Sowahc': '', u'Ripk4': '', u'Hnrnph2': 'Nucleus', u'Lhpp': 'Cytoplasm;Nucleus', u'NISCH': '', u'Prkag2': '', u'TGM2': '', u'Dnmt3a': 'Cytoplasm;Nucleus', u'Snca': 'Cell junction;Cytoplasm;Membrane;Nucleus;Secreted;Synapse', u'Tars': 'Cytoplasm', u'Apob': 'Chylomicron;Cytoplasm;LDL;Secreted;VLDL', u'Actb': 'Cytoplasm;Cytoskeleton', u'Tmprss11d': 'Cell membrane;Membrane;Secreted', u'RGD1563307': '', u'Ggct': '', u'Fut7': 'Golgi apparatus', u'Tcp1': 'Cytoplasm;Cytoskeleton', u'Fut4': 'Golgi apparatus;Membrane', u'Dnajc10': 'Endoplasmic reticulum', u'Gnai2': 'Cell membrane;Cytoplasm;Cytoskeleton;Membrane', u'Gnai3': 'Cell membrane;Cytoplasm;Cytoskeleton;Membrane', u'Eef2': 'Cytoplasm;Nucleus', u'Eif2s2': '', u'tbp': '', u'Slc2a4': 'Cell membrane;Cytoplasm;Membrane', u'Slc2a2': 'Membrane', u'Slc2a1': 'Cell membrane;Membrane', u'Rpl13a': '', u'Yars': 'Cytoplasm', u'Cd63': 'Cell membrane;Endosome;Lysosome;Membrane;Secreted', u'Ctnna1': '', u'Papss1': '', u'Plcb1': 'Cytoplasm;Membrane;Nucleus', u'Papss2': '', u'Plcb4': '', u'Bves': 'Cell junction;Cell membrane;Membrane;Tight junction', u'Hras': 'Cell membrane;Golgi apparatus;Membrane', u'Egfr': 'Membrane', u'Impdh2': 'Cytoplasm;Nucleus', u'Cacng8': 'Cell junction;Cell membrane;Membrane;Postsynaptic cell membrane;Synapse', u'LOC298795': '', u'Rab2b': '', u'Lamc3': '', u'Lamc1': '', u'Gas6': 'Secreted', u'Rac2': '', u'Rac1': 'Cell membrane;Cell projection;Cytoplasm;Membrane', u'mlc1': '', u'Angptl3': '', u'Txndc12': 'Endoplasmic reticulum', u'Mrpl32': '', u'p4hb': '', u'Acer1': 'Membrane', u'Cplx1': 'Cytoplasm', u'Skint4': '', u'Src': 'Cell membrane;Cytoplasm;Cytoskeleton;Membrane;Mitochondrion;Mitochondrion inner membrane;Nucleus', u'Ndufa13': 'Membrane', u'ENSRNOG00000033300': '', u'Srm': '', u'Api5': '', u'Hgf': '', u'stat3': '', u'Agbl1': '', u'Kcnma1': 'Cell membrane;Endoplasmic reticulum;Membrane', u'Kifc1': 'Cytoplasm;Cytoskeleton;Endosome;Microtubule;Nucleus', u'RT1-Da': 'Membrane;MHC II', u'Sdha': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'Tmem199': 'Cytoplasmic vesicle;Endoplasmic reticulum;Membrane', u'Sdhb': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'Tra2b': 'Nucleus', u'Cd74': 'Membrane', u'Mfsd8': 'Membrane', u'dpysl2': '', u'Atp6ap1': 'Endoplasmic reticulum;Membrane', u'Atp6ap2': 'Membrane', u'ENSRNOG00000050913': '', u'Moxd1': '', u'dbn1': '', u'Gucy2c': 'Cell membrane;Endoplasmic reticulum;Membrane', u'atp1a2': '', u'Atp6v0b': 'Membrane', u'Atp6v0c': 'Membrane;Vacuole', u'Cetn2': '', u'Nrxn1': 'Cell junction;Cell membrane;Membrane;Synapse', u'Ahsa1': '', u'Ahsa2': '', u'ENSRNOG00000017095': '', u'dab2': '', u'c2': '', u'ENSRNOG00000013082': '', u'Sytl5': 'Membrane', u'Sytl1': '', u'map3k3': '', u'ENSRNOG00000000656': '', u'Elovl1': 'Endoplasmic reticulum', u'Dctn2': 'Cytoplasm;Cytoskeleton;Dynein;Membrane;Microtubule', u'Ssb': 'Nucleus', u'Dctn6': '', u'Dctn5': '', u'tnk2': '', u'Casp4': '', u'Thap4': '', u'Casp3': 'Cytoplasm', u'Rpsa': 'Cell membrane;Cytoplasm;Membrane;Nucleus', u'Gnb4': '', u'Gnb1': '', u'Gnb3': '', u'Gnb2': 'Cytoplasm', u'ank2': '', u'WASL': '', u'Map1a': 'Cytoplasm;Cytoskeleton;Microtubule', u'Map1b': 'Cell junction;Cell projection;Cytoplasm;Cytoskeleton;Microtubule;Synapse', u'pdlim2': '', u'Tuba1c': 'Cytoplasm;Cytoskeleton;Microtubule', u'Tuba1b': 'Cytoplasm;Cytoskeleton;Microtubule', u'Tuba1a': 'Cytoplasm;Cytoskeleton;Microtubule', u'Cdc37': 'Cytoplasm', u'Cpn2': '', u'Hexdc': '', u'itgb1': '', u'Akap5': 'Membrane', u'dao': '', u'Rps6': '', u'Rps5': '', u'Rps3': 'Cytoplasm;Cytoskeleton;Membrane;Mitochondrion;Mitochondrion inner membrane;Nucleus', u'Brd9': '', u'Phlpp2': '', u'Phlpp1': 'Cell membrane;Cytoplasm;Membrane;Nucleus', u'Timm44': 'Membrane;Mitochondrion;Mitochondrion inner membrane', u'Rps9': 'Cytoplasm', u'Rps8': 'Cytoplasm;Membrane', u'Chia': 'Cytoplasm;Secreted', u'ENSRNOG00000030448': '', u'Lamp1': 'Cell membrane;Endosome;Lysosome;Membrane', u'MAPT': '', u'BAG2': '', u'Tm6sf2': 'Endoplasmic reticulum;Membrane', u'P4hb': 'Cell membrane;Endoplasmic reticulum;Membrane', u'Gab1': '', u'Rsu1': '', u'Zfand2b': 'Endoplasmic reticulum;Membrane', u'Itgb1': 'Cell junction;Cell membrane;Cell projection;Endosome;Membrane', u'Itgb3': 'Membrane', u'Itgb2': 'Membrane', u'Itgb5': '', u'Fam108b1': '', u'Itgb6': 'Cell junction;Membrane', u'Arl8b': 'Cytoplasm;Cytoskeleton;Endosome;Lysosome;Membrane', u'Arl8a': '', u'Fau': '', u'Hsp90ab1': 'Cell membrane;Cytoplasm;Membrane;Nucleus;Secreted', u'Kif5a': 'Cytoplasm;Cytoskeleton;Microtubule', u'Ogdh': 'Mitochondrion;Nucleus', u'Pomc': 'Secreted', u'Ppbp': 'Secreted', u'ENSRNOG00000038248': '', u'Tor2a': 'Secreted', u'Cav1': 'Cell membrane;Golgi apparatus;Membrane', u'Rad23b': 'Cytoplasm;Nucleus;Proteasome', u'Ppp2ca': 'Centromere;Chromosome;Cytoplasm;Cytoskeleton;Nucleus', u'Tac1': 'Secreted', u'Scarb1': 'Cell membrane;Membrane', u'Ppp2cb': 'Centromere;Chromosome;Cytoplasm;Cytoskeleton;Nucleus', u'RGD1562558': '', u'Gpm6a': 'Cell membrane;Cell projection;Membrane', u'Naga': 'Lysosome', u'A4galt': 'Golgi apparatus;Membrane', u'Nagk': '', u'Mterfd1': '', u'Tbce': 'Cytoplasm;Cytoskeleton', u'Tbcd': '', u'Txndc8': 'Cytoplasm;Golgi apparatus', u'Taf9b': 'Nucleus', u'ENSRNOG00000023285': '', u'Txndc2': 'Cytoplasm', u'Txndc5': '', u'Pga5': '', u'Cct8l1': 'Cytoplasm', u'Ttc12': '', u'Kdelr2': 'Endoplasmic reticulum;Membrane', u'Kdelr3': 'Endoplasmic reticulum', u'Lmbrd1': 'Lysosome;Membrane', u'Kdelr1': 'Cytoplasmic vesicle;Endoplasmic reticulum;Membrane', u'Idh3a': 'Mitochondrion', u'Lrig3': 'Membrane', u'Idh3g': 'Mitochondrion', u'Slc44a2': 'Membrane', u'Prkcd': 'Cytoplasm;Membrane;Nucleus', u'Taf4a': '', u'Cck': 'Secreted', u'Sipa1': '', u'Wnt7a': 'Extracellular matrix', u'Wnt7b': 'Extracellular matrix', u'rara': '', u'Mrpl27': '', u'Idh3B': 'Mitochondrion', u'ENSRNOG00000045855': '', u'Mrpl23': 'Mitochondrion', u'tmed9': '', u'Galc': '', u'Galm': 'Cytoplasm', u'RGD1564469': '', u'itpr1': '', u'Erlec1': '', u'Epn1': 'Cell membrane;Coated pit;Cytoplasm;Membrane;Nucleus', u'Ecm1': 'Extracellular matrix;Secreted', u'Epn2': 'Cytoplasm', u'TMED2': '', u'Crnkl1': 'Nucleus;Spliceosome', u'Ranbp1': '', u'mep1b': '', u'Ranbp2': '', u'Ppm1e': 'Cytoplasm;Nucleus', u'Rpl6': 'Cytoplasm;Endoplasmic reticulum', u'Rpl7': '', u'Rpl4': '', u'Des': 'Cell membrane;Cytoplasm;Intermediate filament;Membrane;Nucleus', u'Rpl3': 'Cytoplasm;Nucleus', u'Glp1r': 'Cell membrane;Membrane', u'Mecr': 'Cytoplasm;Mitochondrion;Nucleus', u'Aifm1': 'Cytoplasm;Membrane;Mitochondrion;Mitochondrion inner membrane;Nucleus', u'RGS16': '', u'RGS17': '', u'RGS14': '', u'Tek': 'Membrane', u'dmd': '', u'RGS10': '', u'Atp6v1e1': '', u'RGS18': '', u'Atp6v1e2': '', u'Itih3': 'Secreted', u'RT1.Ha': 'Membrane;MHC II', u'Psmd13': 'Proteasome', u'Psmd12': 'Proteasome', u'Stfa3': '', u'Suclg1': 'Mitochondrion', u'Thnsl1': '', u'ENSRNOG00000047365': '', u'Myo16': 'Cytoplasm', u'Kif20a': 'Microtubule', u'Ampd2': '', u'Ace': 'Cell membrane;Cytoplasm;Membrane;Secreted', u'Ugt8': 'Membrane', u'PDIA5': '', u'Atg5': 'Cytoplasm;Membrane', u'PDIA3': '', u'RGD1307222': '', u'ubc': '', u'Pomt1': 'Endoplasmic reticulum;Membrane', u'Pomt2': 'Membrane', u'Rpl29': '', u'Cyb5d2': 'Secreted', u'Rpl26': '', u'gapdh': '', u'Rpl23': '', u'Chst5': 'Membrane'}


#  To use openpxl to read file
if csvpath.endswith('.xls') or csvpath.endswith('.xlsx'):
    import openpyxl
    wb = openpyxl.load_workbook(csvpath)
    ws = wb.active  # choose active sheet
    from openpyxl.styles import Font
    from openpyxl.styles import Color
    from openpyxl.styles import colors
    red = Font(color=colors.RED)  # provides text for font
    all_data = [['' for i in range(ws.max_column)] for j in range(ws.max_row)]
    rowNumber = -1
    for row in ws.rows:
        rowNumber += 1
        columnNumber = -1
        for column in row:
            columnNumber += 1
            if column.value == None:
                continue
            if type(column.value) == str:
                all_data[rowNumber][columnNumber] = column.value.encode('UTF-8', errors='replace') # I put the encode() in to try to replace weird characters in the original spreadsheet. It doesn't seem to work. FIX
                continue
            all_data[rowNumber][columnNumber] = column.value

# possibly use Pandas:
# import pandas
# df = pandas.DataFrame(ws.values)  # makes it into a table?


# 20180918 BB Tried this but didn't work. Remember Shruti had problems with it too. Pandas doesn't like mixed data types or something
# =============================================================================
# def convertExcelToCSV(filePath):
#     print('convertExcelToCSV')
#     import pandas as pd
#     data_xls = pd.read_excel(filePath, index_col=None)
#     data_xls.to_csv(csvpath, encoding='utf-8')
# #
# # initialPath = raw_input("What is the file path (with extension)? ")
# if csvpath.endswith('.xls') or csvpath.endswith('.xlsx'):
#     convertExcelToCSV(csvpath)
# =============================================================================
#csvpath_BB = "/Users/mac/Dropbox (Scripps Research)/ISG_Goodsell/ISG_Goodsell_BB.csv"

if not os.path.isdir(model_dir + 'PDB'):
    print('making PDB directory')
    os.mkdir(model_dir + 'PDB')

if csvpath.endswith('.csv'):
    with open(csvpath, 'rU') as csvfile:  # need to open the file in Universal mode so it can read Mac Excel output .csv
        spamreader = csv.reader(csvfile)
        for row in spamreader:
            all_data.append(row)


class MyHTMLParser(HTMLParser):
    #
    # def __init__(self, html, lookforTag="a", lookforData="[Show]", lookforAttr=None, gatherData=False):
    #     self.lookforTag = lookforTag
    #     self.lookforData = lookforData
    #     self.lookforAttr = lookforAttr
    #     self.gatherData = gatherData
    #     self.tag_stack = False
    #     self.tag_attr = []
    #     self.stored = []
    #     self.stored_attr = []
    #     self.feed(html)

    def extract_p_contents(self, html, lookforTag="a", lookforData="[Show]", lookforAttr=None, gatherData=False):
        self.lookforTag = lookforTag
        self.lookforData = lookforData
        self.lookforAttr = lookforAttr
        self.gatherData = gatherData
        self.tag_stack = False
        self.tag_attr = []
        self.stored = []
        self.stored_attr = []
        self.feed(html)

    def handle_starttag(self, tag, attrs):
        if tag == self.lookforTag:
            self.tag_stack = True
            self.tag_attr = []
            # print "Encountered the beginning of a %s tag" % tag
            # print attrs
            if len(attrs):
                if self.lookforAttr is not None:
                    for atr in self.lookforAttr:
                        for attr in attrs:
                            if attr[0] == atr:
                                self.tag_attr.append(attr[1])
                else:
                    self.tag_attr = attrs[0][1]

    def handle_endtag(self, tag):
        self.tag_stack = False
        # self.tag_attr = None
        pass

    #        print "Encountered the end of a %s tag" % tag

    def handle_data(self, data):
        if self.tag_stack:
            if self.lookforData is None:
                if self.gatherData:
                    self.stored.append(data)
                else:
                    if len(self.tag_attr):
                        self.stored.append(self.tag_attr)
            else:
                if data == self.lookforData:
                    # print "Encountered data %s" % data
                    # print self.tag_attr
                    self.stored.append(self.tag_attr)



# wb.save(str(model_dir + csvname + '_test1.xlsx'))  # save

taxid = '10116' # default is RAT
# find TAXID
for x in range(0,len(all_data)):
    value = all_data[x][0]
    if value[0:6] == 'TAXID=':
        taxid = value[6:]
        print('TAXID=' + taxid)
        break

# get string to represent taxid in searches
# taxname = 'RAT'  # temporary - replace based on TAXID

taxlong = ''
url = 'https://www.uniprot.org/uniprot/?query=organism%3A' + taxid + '&sort=score'
response = urllib2.urlopen(url)
response_text = response.read()
r = response_text
cursorstring = str('/taxonomy/' + str(taxid) + '">')
cursor = r.find(cursorstring) + len(cursorstring)
if cursor == len(cursorstring)-1:
    taxlong = 'not found'
endcursor = r.find('<', cursor)
if endcursor == -1:
    taxlong = 'error'
if taxlong == '':
    taxlong = r[cursor:endcursor]
print('taxlong = ' + taxlong)

# Localizations are places where given ingredients are found
targetLocalizations = ['Secreted', 'Cytoplasmic vesicle']  #Synaptosome? Peroxisome? Cytoskeleton? Cell membrane? Endosome?
neutralLocalizations = ['Endoplasmic reticulum', 'Membrane', 'Endosome', 'Cell membrane', 'Golgi apparatus',
                        'Peroxisome', '', 'Lysosome', 'Amyloid', 'Synapse', 'Synaptosome', 'Surface film',
                        'Host cell membrane', 'Host membrane', 'Thick filament']
contaminantLocalizations = ['Nucleus', 'Spliceosome','Signalosome', 'Telomere',
                            'Mitochondrion', 'Mitochondrion inner membrane', 'Mitochondrion outer membrane',
                            'Cytoplasm', 'Cytoskeleton', 'Extracellular matrix', 'Mitochondrion nucleoid',
                            'Basement membrane', 'Cell projection', 'Primosome',
                            'Cell junction', 'Cilium', 'Microtubule',
                            'Postsynaptic cell membrane', 'Dynein', 'Centromere', 'Chromosome', 'Tight junction',
                            'Sarcoplasmic reticulum', 'Coated pit', 'Nuclear pore complex', 'Proteasome', 'HDL',
                            'Intermediate filament', 'Vacuole', 'CF(0)', 'Flagellum', 'Lipid droplet', 'Microsome',
                            'Kinetochore', 'Chylomicron', 'VLDL', 'LDL', 'MHC II', 'MHC I', 'Gap junction',
                            'Nucleosome core', 'Virion', 'CF(1)', 'DNA-directed RNA polymerase', 'Keratin',
                            'Membrane attack complex', 'Target cell membrane', 'Target membrane',
                            'Signal recognition particle', 'Host cytoplasm', 'Exosome', 'Host nucleus',
                            'Capsid protein', 'Host endosome', 'Inflammasome', 'Host mitochondrion']
newLocalizations = []

# Locations are places where interacting proteins are found
targetLocations = ['Secreted', 'Cytoplasmic vesicle']  #Synaptosome? Peroxisome? Cytoskeleton? Cell membrane? Endosome?
neutralLocations = ['Membrane', 'Endosome', 'Cell membrane', 'Cytoplasm', 'Cytoskeleton', 'Extracellular matrix',
                    'Basement membrane', 'Golgi apparatus', 'Peroxisome', 'Cell projection', 'Lysosome',
                    'Cell junction', 'Synapse', 'Synaptosome', 'Cilium', 'Microtubule', 'Surface film', 'Host cell membrane', 'Host membrane', 'Thick filament', 'Primosome',
                    'Postsynaptic cell membrane', 'Dynein', 'Centromere', 'Chromosome', 'Tight junction',
                    'Sarcoplasmic reticulum', 'Coated pit', 'Nuclear pore complex', 'Proteasome', 'HDL',
                    'Intermediate filament', 'Vacuole', 'CF(0)', 'Flagellum', 'Lipid droplet', 'Microsome',
                    'Kinetochore', 'Chylomicron', 'VLDL', 'LDL', 'MHC II', 'MHC I', 'Gap junction', 'Nucleosome core',
                    'Virion', 'CF(1)', 'DNA-directed RNA polymerase', 'Amyloid', 'Keratin','Signalosome', 'Telomere',
                    'Membrane attack complex', 'Target cell membrane', 'Target membrane', 'Signal recognition particle',
                    'Host cytoplasm', 'Exosome', 'Host nucleus', 'Capsid protein', 'Host endosome', 'Inflammasome',
                    'Host mitochondrion']
contaminantLocations = ['Endoplasmic reticulum', 'Nucleus', 'Spliceosome', 'Mitochondrion nucleoid',
                         'Mitochondrion', 'Mitochondrion inner membrane', 'Mitochondrion outer membrane']
newLocations = []



# look for headers and assign column numbers
for x in range(0,len(all_data)):
    value = all_data[x][0]
    if value == 'HEADERS':
        headersrow = x
        print('headersrow = ' + str(headersrow + 1))
        break

default_headers = ['HEADERS', 'INCLUDE', 'CONFIDENCE', 'ACCESSION', 'UNIPROT_ID', 'ORGANISM', 'GENE',
                   'UNIPROT_NAME', 'UNIPARC', 'NAME', 'SEQUENCE_ORIGINAL', 'MODIFIED SEQUENCE','FOUNDPDBS', 'chosenPDB',
                   'TOP_PDB', 'OPM', 'CLEANOPM', 'PRINCIPAL_VECTOR', 'OFFSET', 'JITTER_MAX', 'COLOR', 'MW',
                   'LENGTH', 'notes', 'LOCALIZATION', 'MEMBRANE', 'COFACTORS', 'LIGANDS', 'INTERACTIONS_RGD',
                   'INTERACTIONS_UNIPROT', 'INTERACTIONS_STRINGDB', 'INTERACTIONS_BIOGRID', 'INTERACTIONS_INTACT']
startingHeaders = all_data[headersrow]
for x in range(len(startingHeaders)):
    if startingHeaders[x] not in all_data[headersrow]:
        ws.insert_cols(x+1)
        for y in range(len(all_data)):
            all_data[y].insert(x, '')
            ws.cell(y+1,x+1,'')
        all_data[headersrow][x] = startingHeaders[x]

print('HEADERS')
print(all_data[headersrow])
# wb.save(str(model_dir + csvname + '_test2.xlsx'))  # save

# possibly only add headers when needed? change values as new headers are added?


# headers = {'test': 'headers test works'}
headers = {}
for num in range(len(all_data[headersrow])):
    headers[all_data[headersrow][num]] = num  # This establishes a dictionary with the header names in it.
    # After this, columns can be indicated with e.g. "name = all_data[x][headers['NAME']]".
    # The headers must be correctly labeled.


# acc2seq will take an accession number as input and return the sequence of that molecule


if os.path.isfile(model_dir + csvname + '_inprogress.csv'):
    print('inprogress file exists')


def acc2fasta(accession):  # I did this using Entrez, but why? Just makes it dependent. Maybe change.
    print('acc2seq ' + accession)
    handle = Entrez.efetch(db="protein", id=accession, rettype="fasta")
    fasta = handle.read()
    return fasta


def acc2seq(accession):  # I did this using Entrez, but why? Just makes it dependent. Maybe change.
    print('acc2seq ' + accession)
    try:
        acc = accession.split(';')
    except:
        acc = accession
    print ('acc2seq (' + acc + ')')
    handle = Entrez.efetch(db="protein", id=acc, rettype="fasta")
    response = handle.read()
    response_split = response.split('\n')
    header, sequence = response_split[0], response_split[1:]
    sequence = ''.join(sequence)
    return sequence

# Returns the subcellular location of a protein by query- strips the query of any unneccesary information
def uniprot2localizationSV(uniprot):
    print('uniprot2localization ' + uniprot)
    allLocations = []
    finalString = ""
    uniprotFile = urllib.urlopen('https://www.uniprot.org/uniprot/?query=accession:' + uniprot + '&columns=comment(SUBCELLULAR LOCATION)&format=tab')
    new = uniprotFile.read().strip('Subcellular location [CC]')
    updated = new.replace('.', '; ')
    listOfLocations = updated.split('; ')
    for i in listOfLocations:
        # Strips the query information of the publication information
        if '{' in i:
            toAdd = i.replace(',', ':')
            allLocations.append(toAdd[:toAdd.index('{')].strip())
        else:
            toAdd = i.replace(',', ':')
            allLocations.append(toAdd.strip())
    # Deletes all redundant entries from those found by the query
    finalList = list(set(allLocations))
    for i in finalList:
        if len(i) > 0:
            finalString += i + ", "
    # Returns the list of subcellular localatiobs
    return finalString[:len(finalString)-2]


def uniprot2localization(uniprot):
    if uniprot == 'None' or uniprot == '':
        return ''
    print ('uniprot2localization (' + uniprot + ')')
    url = 'http://www.uniprot.org/uniprot/' + uniprot
    response = urllib2.urlopen(url)
    response_text = response.read()
    r = response_text
    cursorstring = 'Keywords - Cellular component'
    cursor = r.find(cursorstring) + len(cursorstring)
    if cursor == len(cursorstring)-1:
        return ''
#    while r.find('cursor = r.find('">', cursor)
    endcursor = r.find('class=', cursor)
    localization = r[cursor:endcursor]
    cursor = 0
    locations = []
    while localization.find('">', cursor) != -1:
        cursor = localization.find('">', cursor)+2
        endcursor = localization.find('</a>', cursor)
        location = localization[cursor:endcursor]
        locations.append(location)
        cursor = endcursor
    locationstring = ';'.join(locations)
    print ('locations = ' + locationstring)
    return locationstring


def gi2acc(gi):
    print('gi2acc(' + gi + ')')
    url = 'https://www.ncbi.nlm.nih.gov/protein/' + gi
    response = urllib2.urlopen(url)
    r = response.read()
    cursorstring = 'BlastSearch&amp;QUERY='
    cursor = r.find(cursorstring) + len(cursorstring)
    if cursor == len(cursorstring)-1:
        return ''
    endcursor = r.find('"', cursor)
    if endcursor == -1:
        endcursor = r.find('</p>', cursor)
    accession = r[cursor:endcursor]
    print('    acc = ' + accession + '.')
    return accession


def acc2uniparc(accession):
    print('acc2uniparc(' + accession + ')')
    try:
        accs = accession.split(';')
        acc = accs[0]
        print('using ' + str(acc))
    except:
        acc = accession
    try:
        acc, vers = acc.split('.')
        print('using ' + str(acc))
    except:
        pass
    url = 'http://www.uniprot.org/uniparc/?query=' + acc + '&sort=score'
    response = urllib2.urlopen(url)
    r = response.read()
    cursor = r.find('tr id="UPI') + 7
    if cursor == 6:
        return ''
    endcursor = r.find('"', cursor)
    print(cursor, endcursor)
    uniparc = r[cursor:endcursor]
    print('    uniparc = ' + uniparc + '.')
    return uniparc


# def ipi2uniprot(ipi):
#     print('ipi2uniprot(' + ipi + ')')
#     url = 'http://www.uniprot.org/uniparc/?query=' + ipi + '&sort=score'
#     response = urllib2.urlopen(url)
#     response_text = response.read()
#     r = response_text
#     cursor = 0
#     cursor = r.find('<a href="/uniprot/', cursor) + 18
#     if cursor == 17:
#         return ''
#     endcursor = r.find('?version', cursor, cursor+20)
#     if endcursor == -1:
#         endcursor = r.find('">', cursor, cursor+20)
#     uniprot = r[cursor:endcursor]
#     print ('    uniprot = ' + uniprot)
#     r = urllib.urlopen("http://www.uniprot.org/uniprot/" + uniprot + ".xml").read()
#     if r:
#         print('Uniprot is valid')
#     else:
#         print('Uniprot is not valid')
#         uniprot = 'complicated'
#     return uniprot


# def ipi2uniparc(ipi):
#     print('ipi2uniparc(' + ipi + ')')
#     url = 'http://www.uniprot.org/uniparc/?query=' + ipi + '&sort=score'
#     response = urllib2.urlopen(url)
#     response_text = response.read()
#     r = response_text
#     cursor = r.find('tr id="UPI') + 7
#     if cursor == -1:
#         return ''
#     endcursor = r.find('"', cursor)
#     uniparc = r[cursor:endcursor]
#     print('    uniparc = ' + uniparc)
#     return uniparc


# def ipi2uniparcAndUniprot(ipi):
#     print('ipi2uniparc(' + ipi + ')')
#     url = 'http://www.uniprot.org/uniparc/?query=' + ipi + '&sort=score'
#     response = urllib2.urlopen(url)
#     response_text = response.read()
#     r = response_text
#     records = r.split('entryID')
#     for record in records:
#         if taxlong in record and '/uniparc/' in record and '/uniprot/' in record:
#             uniparc = (record.split('/uniparc/')[1]).split('"')[0]
#             uniprot = (record.split('/uniprot/')[1]).split('"')[0]
#         elif taxlong in record and '/uniparc/' in record:
#             return uniparc, uniprot
#     return '',''


def ipi2uniparc(ipi):
    print('ipi2uniparc(' + ipi + ')')
    url = 'http://www.uniprot.org/uniparc/?query=' + ipi + '&sort=score'
    response = urllib2.urlopen(url)
    response_text = response.read()
    r = response_text
    records = r.split('entryID')
    for record in records:
        record = record.strip()
        if taxlong in record and '/uniparc/' in record:
            return (record.split('/uniparc/')[1]).split('"')[0]
    return ''


def acc2uniprot(acc):
    print('acc2uniprot(' + acc + ')')
    url = 'https://www.uniprot.org/uniprot/?query=' + acc + '&sort=score'
    response = urllib2.urlopen(url)
    response_text = response.read()
    r = response_text
    records = r.split('entryID')
    for record in records:
        if taxlong in record and '/uniprot/' in record:
            uniprot = (record.split('/uniprot/')[1]).split('"')[0]
            return uniprot
    return ''


def pdb2organism(pdb):
    url = 'https://www.rcsb.org/structure/'

# gets protein gene, name, organism, and review info from Uniprot
def getOrganism(uniprot):
    print('getOrganism ' + uniprot)
    uniprotFile = urllib.urlopen("http://www.uniprot.org/uniprot/" + uniprot + ".txt")
    for line in uniprotFile:
        if line[0:2] == 'OS':
            organism = (line.split('OS')[1]).split('.')[0]
            break
    return organism.strip()


def getGene(uniprot):
    print('getGene ' + uniprot)
    uniprotFile = urllib.urlopen("http://www.uniprot.org/uniprot/" + uniprot + ".txt")
    for line in uniprotFile:
        if line[0:2] == 'GN':
            return ((line.split('=')[1]).split(';')[0]).split(' {')[0]


def getUniprotName(uniprot):
    print('getUniprotName ' + uniprot)
    uniprotFile = urllib.urlopen("http://www.uniprot.org/uniprot/" + uniprot + ".txt")
    for line in uniprotFile:
        if "Name: Full=" in line:
            return ((line.split('Full=')[1]).split(';')[0]).split(' {')[0]


def getAnnotationScore(uniprot):
    print('getAnnotationScore ' + uniprot)
    uniprotFile = urllib.urlopen("http://www.uniprot.org/uniprot/" + uniprot)
    for line in uniprotFile:
        if "<p>Annotation score:" in line:
            return (line.split('<p>Annotation score:')[1]).split(' ')[0]


def getReviewed(uniprot):
    print('getReviewed ' + uniprot)
    uniprotFile = urllib.urlopen("http://www.uniprot.org/uniprot/" + uniprot + ".txt")
    for line in uniprotFile:
        if "ID  " in line:
            if 'Reviewed' in line:
                return 'X'
            break


# def getUniref50(uniprot):
#     print('getUniref50 ' + uniprot)
#     uniprotFile = urllib.urlopen("http://www.uniprot.org/uniprot/" + uniprot)
#     for line in uniprotFile:
#         if "UniRef50_" in line:
#             return (line.split('UniRef50_')[1]).split('"')[0]
#     return('N/A')
#
#
# def getUniref90(uniprot):
#     print('getUniref90 ' + uniprot)
#     uniprotFile = urllib.urlopen("http://www.uniprot.org/uniprot/" + uniprot)
#     for line in uniprotFile:
#         if "UniRef90_" in line:
#             return (line.split('UniRef90_')[1]).split('"')[0]
#     return('N/A')
#
#
# def getUniref100(uniprot):
#     print('getUniref100 ' + uniprot)
#     uniprotFile = urllib.urlopen("http://www.uniprot.org/uniprot/" + uniprot)
#     for line in uniprotFile:
#         if "UniRef100_" in line:
#             return (line.split('UniRef100_')[1]).split('"')[0]
#     return('N/A')
#
#
#
# def getUniRef50Top(uniRef):
#     print('getUniRef50Top ' + uniRef)
#     uniprotFile = urllib.urlopen("https://www.uniprot.org/uniref/UniRef50_" + uniRef + ".xml")
#     for line in uniprotFile:
#         if "_RAT" in line:
#             nextline = uniprotFile.readline()
#             if "UniProtKB accession" in nextline:
#                 return (nextline.split('"UniProtKB accession" value="')[1]).split('"')[0]
#     return('N/A')
#
#
# def getUniRef90Top(uniRef):
#     print('getUniRef90Top ' + uniRef)
#     uniprotFile = urllib.urlopen("https://www.uniprot.org/uniref/UniRef90_" + uniRef + ".xml")
#     for line in uniprotFile:
#         if "_RAT" in line:
#             nextline = uniprotFile.readline()
#             if "UniProtKB accession" in nextline:
#                 return (nextline.split('"UniProtKB accession" value="')[1]).split('"')[0]
#     return('N/A')
#
#
# def getUniRef100Top(uniRef):
#     print('getUniRef100Top ' + uniRef)
#     uniprotFile = urllib.urlopen("https://www.uniprot.org/uniref/UniRef100_" + uniRef + ".xml")
#     for line in uniprotFile:
#         if "_RAT" in line:
#             nextline = uniprotFile.readline()
#             if "UniProtKB accession" in nextline:
#                 return (nextline.split('"UniProtKB accession" value="')[1]).split('"')[0]
#     return('N/A')


def getBestUniprotFromGene(gene):
    print('getBestUniprotFromGene ' + gene)
    genea=gene.lower()
    r = urllib.urlopen("https://www.uniprot.org/uniprot/?query=" + gene + "+" + taxid + "&sort=score").read()
    records = r.split('entryID')
    for record in records[1:]:
        if str(taxid) in record and '/uniprot/' in record and str('>' + genea + '<') in record.lower():
            uniprot = (record.split('/uniprot/')[1]).split('"')[0]
            print(str(genea) + ' uniprot = ' + str(uniprot))
            return uniprot
    return 'None'


#Checks to make sure that the uniprot entry is for the rattus norvegicus species. Returns boolean
def checkOrganism(uniprot):
    return True
    print('checkOrganism ' + uniprot)
    urllib.urlretrieve("http://www.uniprot.org/uniprot/" + uniprot + ".xml", model_dir + 'uniprotentry.xml')
    record_iterator = SeqIO.parse(model_dir + 'uniprotentry.xml', 'uniprot-xml')
    try:
        record = next(record_iterator)
    except:
        return False
    if(record.annotations["organism"] == 'Rattus norvegicus (Rat)'):
        print('The uniprot entry is for a rat')
        return True
    print('The uniprot entry is not for a rat')
    return False

# Returns a list of only the similarities between two other lists
def returnSame(li2):
    print('returnSame' + uniprot)
    li1 = ["Magnesium", "Calcium", "Nickel", "Zinc", "ATP-binding", "GTP-binding", "Potassium", "Iron", "Iron-sulfur", "Biotin", "Cadmium", "cAMP", "cGMP", "Chitin-binding", "Chromophore", "Cobalt", "Copper", "Flavoprotein", "Folate-binding", "Hemoglobin-binding", "Hyaluronic acid", "Lectin", "Lipid-binding", "Lithium", "Manganese", "Mercury", "Molybdenum", "FAD", "NAD", "NADP", "Nickel", "Pigment", "Plastoquinone", "PQQ", "Pyridoxal phosphate", "Retinol-binding", "Schiff base", "Selenium", "Chloride", "Sodium", "Lead", "2Fe-2S", "3Fe-4S", "4Fe-4S", "Bacteriochlorophyll", "Bile pigment", "cAMP-binding", "cGMP-binding", "Chitin-binding", "Chromophere", "Cobalamin"]
    return (list(set(li1)- (set(li1) - set(li2))))

# If there's a cofactor, calls returnCofactor function as well as returnLigand function. If not, only calls returnLigands function.
def findCofactorsAndLigands(uniprot):
    print('findCofactorsAndLigands ' + uniprot)
    cofactorsAndLigands = ["", ""]
    presenceOfCofactor = False
    uniprotFile = urllib.urlopen("http://www.uniprot.org/uniprot/" + uniprot + ".txt")
    for line in uniprotFile:
        if 'COFACTOR' in line:
            presenceOfCofactor = True
            break
    if presenceOfCofactor == True:
        cofactorsAndLigands = [returnCofactor(uniprot), returnLigand(uniprot)]
    else:
        cofactorsAndLigands = ["", returnLigand(uniprot)]  # bbq - why did she do it this way?
    return cofactorsAndLigands

# Returns the name of the protein's cofactor
def returnCofactor(uniprot):
    print('returnCofactor ' + uniprot)
    lineNum=0
    lineToPrint=-1
    cofactor = "None"
    uniprotFile = urllib.urlopen("http://www.uniprot.org/uniprot/" + uniprot + ".txt")
    for line in uniprotFile:
        lineNum+=1
        if 'COFACTOR:' in line:
            cofactor = ""
            lineToPrint= lineNum+1
        if lineToPrint==lineNum and 'Name=' in line:
            cofactor += (line.split('Name=')[1]).split(';')[0] + " " #bb stacking commands like this works! probably need ()
    return cofactor.strip()

# Parses xml file for uniprot entry to see whether or not there's a ligand.
def returnLigand(uniprot):
    print('returnLigand ' + uniprot)
    cofactorsAndLigands = "None"
    record_iterator = SeqIO.parse(model_dir + 'uniprotentry' + uniprot + '.xml', 'uniprot-xml')
    record = next(record_iterator)
    try:
        listOfKeywords= record.annotations['keywords']
        finalList = returnSame(listOfKeywords)
        if (len(finalList) != 0):
            cofactorsAndLigands = '; '.join(finalList)
    except:
        pass
    return cofactorsAndLigands

# Returns the interactions listed by the Rat Gene Database
def rgdDependencies(uniprot):
    print('rgdDependencies' + uniprot)
    # Makes use of HTML parser BeautifulSoup
    try:
        from BeautifulSoup import BeautifulSoup
    except ImportError:
        from bs4 import BeautifulSoup
    interaction = ""
    RGD_ID = 'Not found'
    data = urllib.urlopen("http://www.uniprot.org/uniprot/" + uniprot + ".txt")
    for line in data:
        if 'RGD;' in line:
            lineWithRGD= line
            RGD_ID = (lineWithRGD.split('; '))[1]
            break
    html_content = urllib.urlopen('https://rgd.mcw.edu/rgdweb/report/gene/main.html?id=' + RGD_ID).read()
    parsed_html = BeautifulSoup(html_content, 'lxml')
    importantInfo = parsed_html.find('meta', {'name': 'description'})
    importantString= str(importantInfo)
    try:
        listOfWords = importantString.split('INTERACTS WITH ')
    except:
        pass
    try:
        interaction = (listOfWords[1].split('"'))[0]
    except:
        pass
    return(interaction)

# Returns any interactions listed explicitly on the uniprot website (other than cofactors and lignads)
def uniprotInteractions(uniprot):
    print('uniprotInteractions ' + uniprot)
    subunits = ""
    finalReturn = []
    listOfInteractors = []
    uniprotFile = urllib.urlopen('https://www.uniprot.org/uniprot/?query=accession:' + uniprot + '&columns=interactor&format=tab')
    interactors = uniprotFile.read()
    eachGene = ""
    record_iterator = SeqIO.parse(model_dir + 'uniprotentry' + uniprot + '.xml', 'uniprot-xml')  # Why access this?
    record = next(record_iterator)  # What is purpose of next()?
    try:
        subunits = record.annotations['comment_subunit']
    except:
        pass
    # (subunits.strip('[\'')).strip('\']')
    subunits = '; '.join(subunits)
    new = interactors.strip('Interacts')
    new2 = new.strip(' with\n')
    if new2:
        listOfInteractors = new2.split('; ')
    for i in listOfInteractors:
        if i != 'Itself':
            urllib.urlretrieve("http://www.uniprot.org/uniprot/" + i + ".xml", model_dir + 'uniprotentry' + i + '.xml')
            record_iterator = SeqIO.parse(model_dir + 'uniprotentry' + i + '.xml', 'uniprot-xml')
            record = next(record_iterator)
            eachGene += record.annotations['gene_name_primary'] + " "
    if eachGene:
        finalReturn = ['',eachGene]
    if subunits:
        finalReturn = [subunits, eachGene]
    return finalReturn

# Finds all of the interactors listed on the String database for a given protein
def stringdbInteractors(uniprot):
    print('stringdbInteractions ' + uniprot)
    num = 0
    data = urllib.urlopen("https://string-db.org/api/tsv/interaction_partners?identifiers=" + uniprot + "&limit=10")
    stringInteractors = ""
    finalReturn = ""
    # Reads first line
    line = data.readline()
    # While there are still lines to read and you haven't recieved an error. The max number of interactors is currently set as 10.
    while line and num <11:
        if(num>0):
            try:
                l = line.strip().split("\t")
                stringInteractors += l[3] +" "
            except IndexError:
                num+=20  #bb 20 is just random # > 11
        line = data.readline()
        num+=1
    if len(stringInteractors) > 0:
        finalReturn +=stringInteractors
    return finalReturn

# Returns the interactors listed on the BioGrid Database
def biogridInteractors(uniprot):
    print('biogridInteractions ' + uniprot)
    num = 0
    # First retrieves the name of the protein associated with the given uniprot entry
    record_iterator = SeqIO.parse(model_dir + 'uniprotentry' + uniprot + '.xml', 'uniprot-xml')
    record = next(record_iterator)
    name = str(record.annotations['gene_name_primary']).lower()
    biogridInteractors = ""
    listOfInteractors = []
    # Uses API to get the information for a specific protein
    data = urllib.urlopen("https://webservice.thebiogrid.org/interactions/?additionalIdentifierTypes=SWISS-PROT&geneList=" + uniprot + "&includeInteractors=true&accesskey=dcf7d81c2aeac29c09fc38999ec3477c", context=ssl._create_unverified_context())
    line = data.readline()
    # While there are lines to read and you haven't gotten an error- num is not changed inside the while loop because there is no limit to the number of interactors presented by BioGrid
    while line and num <10:
        try:
            l = line.strip().split("\t")
            if len(l[7]) > 1:
                # Checks the interactor at column 7 and at 8 because of inconsistent formatting in the database
                # If the entry at column 7 is not the protein itself, it will be an interactor that should be added.
                if l[7].lower() != name:
                    listOfInteractors.append(l[7].lower())
                # Otherwise, check to see if there is an entry at column 8 and add that
                elif len(l[8]) >1:
                    listOfInteractors.append(l[8].lower())
        except IndexError:
                num+=20
        line = data.readline()
    # Deletes any reduncancies
    finalList = list(set(listOfInteractors))
    for i in finalList:
        biogridInteractors += i + " "
    return biogridInteractors

# Finds the interactors listed on the IntAct website
def intactInteractors(uniprot):
    print('intactInteractions ' + uniprot)
    # Attempts to import BeautifulSoup. Using BeautifulSoup vs. bs4 depends on the version of python being used
    try:
        from BeautifulSoup import BeautifulSoup
    except ImportError:
       from bs4 import BeautifulSoup
    try:
        gene = getGene(uniprot)
    except:
        gene = ''
    # First retrieves the name of the protein associated with the given uniprot entry
    intactInteractors = ""
    finalReturn = ""
    record_iterator = SeqIO.parse(model_dir + 'uniprotentry' + uniprot + '.xml', 'uniprot-xml')  # problem with this is that it only works if this function is called under certain circumstances
    record = next(record_iterator)
    name = str(record.annotations['gene_name_primary']).lower()
    # print(name)
    # print(gene)
    data = urllib.urlopen("https://www.ebi.ac.uk/intact/query/" + uniprot).read()
    parsed_html = BeautifulSoup(data, 'lxml')
    information = []
    try:
        table = parsed_html.find('tbody', {'id': "mainPanels:resultsTable_data"})
        information = table.findAll('div', {'class' : "ui-dt-c"})
    # This would occur if intact does not have an entry for this protein. Nothing will be done because both intactInteractors and finalReturn are empty and will be returned that way.
    except AttributeError:
        pass
    # Finds all the possible entries on the table that could be an interactor. Weeds through them based on the characteristic length of a protein name
    for i in information:
        # print(i.contents)
        if 5 < len(str(i.contents)) < 60:
            listS = str(i.contents).split('\'')
            # print('listS ' + str(listS))
            # print(listS[1])
            if listS[1].lower() != name and listS[1] != gene:  # don't know why SV used "lower()" here
                intactInteractors += listS[1].split('\\')[0] + " "
                # print intactInteractors
    if len(intactInteractors) > 1:
        finalReturn += intactInteractors
        print('intactInteractors = ' + finalReturn)
    return finalReturn

# Retrieves the ending index of the singal peptide and returns the complete sequence starting at that point, or 'None' if there is none
def cleaveSignalPeptide(uniprot, sequence):
    print('cleaveSignalPeptide ' + uniprot)
    try:
        uniprotFile = urllib.urlopen('https://www.uniprot.org/uniprot/?query=accession:' + uniprot + '&columns=feature(SIGNAL)&format=tab')  # returns ONLY "Signal peptide" info
        signalInfo = uniprotFile.read()
        usefulInfo = signalInfo.strip('Signal peptide\n')  # bb removes 'Signal peptide' text
        listOfInfo = usefulInfo.split(' ')
        newSequence = ""
        if len(listOfInfo) > 1:
            endIndex = listOfInfo[2]
            newSequence = sequence[int(endIndex):]
        else:
            newSequence = sequence
        return newSequence
    except:
        print('failed')

def getSignalPeptide(uniprot):
    print('getSignalPeptide ' + uniprot)
    uniprotFile = urllib.urlopen('https://www.uniprot.org/uniprot/?query=accession:' + uniprot + '&columns=feature(SIGNAL)&format=tab')  # returns ONLY "Signal peptide" info
    signalInfo = uniprotFile.read()
    usefulInfo = signalInfo.strip('Signal peptide\n')  # bb removes 'Signal peptide' text
    listOfInfo = usefulInfo.split(' ')
    newSequence = ""
    if len(listOfInfo) > 1:
        endIndex = listOfInfo[2]
        newSequence = sequence[int(endIndex):]
    return usefulInfo

# Retrieves amino acid sequence from uniprot entry by query
def uniprot2seq(uniprot):
    print('uniprot2seq ' + uniprot)
    sequence = ""
    if uniprot != '' and uniprot != '?' and uniprot != 'complicated':
        uniprotFile1 = urllib.urlopen('https://www.uniprot.org/uniprot/?query=accession:' + uniprot + '&columns=sequence&format=tab')
        # Strips the query information of a header in case no actual sequence is present
        sequence =  uniprotFile1.read().strip('Sequence\n')
    return sequence

def uniparc2seq(uniparc):
    print('uniparc2seq(' + uniparc + ')')
    url = 'http://www.uniprot.org/uniparc/' + uniparc + '.fasta'
    response = urllib2.urlopen(url)
    response_text = response.read()
    return response_text.split('active\n')[1]

def uniparc2uniprot(uniparc):
    print('uniparc2uniprot(' + uniparc + ')')
    url = 'http://www.uniprot.org/uniparc/' + uniparc + '.xml'
    response = urllib2.urlopen(url)
    r = response.read()
    records = r.split('<dbReference type="')
    for record in records:
        if taxid in record and 'UniProtKB' in record:
            return (record.split('id="')[1]).split('"')[0]
        elif taxid in record:
            response = (record.split('id="')[1]).split('"')[0]
            url = 'https://www.uniprot.org/uniprot/?query=' + response + '&sort=score'
            try:
                check = urllib2.urlopen(url).read()
                nextasdf = (check.split('"entryID"><a href="/uniprot/')[1]).split('"')[0]
                print(nextasdf)
                return nextasdf
            except:
                return "None"
            break



def seq2pdb(row, sequence, details=0):
    print('seq2pdb ' + sequence[:10] + '...')
    pdb = length = score = expect = method = identities = positives = gaps = ''
    url = 'https://www.rcsb.org/pdb/rest/getBlastPDB1?sequence=' + sequence + '&eCutOff=10.0&matrix=BLOSUM62'
    for x in range(1,4):
        try:
            response = urllib2.urlopen(url)
            response_text = response.read()
            r = response_text
            break
        except:
            print("   urllib2 failure " + str(x))
            if x == 3:
                print("   skipping")
                return ''
    response_text_split = response_text.split('\n')
    if response_text_split[37][:4] == '  Nu':  # this is what it returns if it doesn't find anything
        print("   no pdbs found")
        return ''
    pdbs = []
    for x in range(0,1):  # set second variable to number of results desired
        foundpdb = response_text_split[37 + x][:4]
        if foundpdb == '</PR':
            break
        pdbs.append(foundpdb)
    # detailed information for each entry looks like this:
    # ><a name = 56222></a>2NSM:1:A|pdbid|entity|chain(s)|sequence
    #           Length = 439
    #
    #  Score =  418 bits (1075), Expect = e-117,   Method: Composition-based stats.
    #  Identities = 211/402 (52%), Positives = 275/402 (68%), Gaps = 19/402 (4%)
    #
    # Query: 49  ISFEYHRYPELREALVSVWLQCTAISRIYTVGRSFEGRELLVIELSDNPGVHEPGEPEFK 108
    # ...
    # ... some have no "Gaps". Not sure if there are other variations.
    if details == 1:
        detailheaders = ['pdb', 'length', 'score', 'expect', 'method', 'identities', 'positives', 'gaps', 'title',
                         'organism', 'deposited', 'global_stoichiometry', 'pseudo_stoichiometry', 'weight', 'atoms',
                         'residues', 'chains', 'OPM']
        for dheader in detailheaders:
            if dheader not in all_data[headersrow]:
                insertcolumn(dheader,len(all_data[headersrow]))
        for num in range(len(all_data[headersrow])):
            headers[all_data[headersrow][num]] = num  # This establishes a new dictionary with the header names in it.
        print(all_data[headersrow])
        print(headers)
        cursor = 0
        for x in range(0, 1):  # return up to the first 10 (11?) results
            cursor = r.find('<a name', cursor)  # takes us to next PDBID
            if cursor == -1:
                break
            endcursor = r.find('</a>', cursor) + 4
            pdb = r[endcursor:endcursor + 4]
            print('\npdb = ' + pdb)
            cursor = r.find('Length', cursor) + 9
            endcursor = r.find('\n', cursor)
            length = r[cursor:endcursor]
            print(length)

            cursor = r.find('Score', cursor) + 8
            endcursor = r.find(',', cursor)
            score = r[cursor:endcursor].strip()
            print(score)

            cursor = r.find('Expect', cursor) + 9
            endcursor = r.find(',', cursor)
            expect = r[cursor:endcursor]
            if expect[0] == 'e':
                expect = '1' + expect
            print(expect)

            cursor = r.find('Method', cursor) + 8
            endcursor = r.find('.', cursor)
            method = r[cursor:endcursor]
            print(method)

            cursor = r.find('Identities', cursor) + 13
            endcursor = r.find(')', cursor) + 1
            identities = r[cursor:endcursor]
            print(identities)

            cursor = r.find('Positives', cursor) + 12
            endcursor = r.find(')', cursor) + 1
            positives = r[cursor:endcursor]
            print(positives)

            if not r.find('Gaps', cursor) == -1:
                cursor = r.find('Gaps', cursor) + 7
                endcursor = r.find('\n', cursor)
                gaps = r[cursor:endcursor]
                print(gaps)

            #  NOW TO GET INFO ON THE PDB FILE
            pdbfile = urllib2.urlopen('https://www.rcsb.org/structure/' + pdb)
            p = pdbfile.read()
            pcursor = p.find('citation_title') + 25
            pendcursor = p.find('"', pcursor)
            title = p[pcursor:pendcursor]
            print(title)

            pcursor = p.find('header_organism', pcursor)
            pcursor = p.find('<a class', pcursor)
            pcursor = p.find('>', pcursor) + 1
            pendcursor = p.find('<', pcursor)
            organism = p[pcursor:pendcursor]
            print(organism)

            pcursor = p.find('header_deposited', pcursor) + 65
            pendcursor = p.find('&', pcursor)
            deposited = p[pcursor:pendcursor]
            print(deposited)

            # pcursor = p.find('Global Stoichiometry') +31
            # pendcursor = p.find('&', pcursor)
            # global_stoichiometry = p[pcursor:pendcursor]
            # pcursor = pendcursor + 41
            # pendcursor = p.find('&', pcursor)
            # global_stoichiometry = global_stoichiometry + ' ' + p[pcursor:pendcursor]
            # print(global_stoichiometry)
            #
            au = ba1 = ''
            sections = p.split('BiologicalUnit')
            for section in sections:
                if "Asymmetric Unit" in section:
                    au = (section.split("Global Stoichiometry</strong>: ")[1]).split(' -&nbsp')[0]
                    print('AU = ' + au)
                if "Biological Assembly 1" in section:
                    ba1 = (section.split("Global Stoichiometry</strong>: ")[1]).split(' -&nbsp')[0]
                    print('BU1 = ' + ba1)
            if ba1:
                global_stoichiometry = 'BA1:' + (section.split("Global Stoichiometry</strong>: ")[1]).split(' -&nbsp')[0]
            else:
                global_stoichiometry = 'AU:' + (section.split("Global Stoichiometry</strong>: ")[1]).split(' -&nbsp')[0]
            print(global_stoichiometry)

            if not p.find('Pseudo Stoichiometry') == -1:
                pcursor = p.find('Pseudo Stoichiometry') +31
                pendcursor = p.find(' -', pcursor)+2
                pseudo_stoichiometry = p[pcursor:pendcursor]
                pcursor = p.find('<span', pendcursor) + 36
                pendcursor = p.find('&', pcursor)
                pseudo_stoichiometry = pseudo_stoichiometry + ' ' + p[pcursor:pendcursor]
                print(pseudo_stoichiometry)
            else:
                pseudo_stoichiometry = ''

            pcursor = p.find('Total Structure Weight') + 24
            pendcursor = p.find('&', pcursor)
            weight = p[pcursor:pendcursor]
            print(weight)

            pcursor = p.find('Atom Count: ', pcursor) + 12
            pendcursor = p.find('&', pcursor)
            atoms = p[pcursor:pendcursor]
            print(atoms)

            pcursor = p.find('Residue Count: ', pcursor) + 15
            pendcursor = p.find('&', pcursor)
            residues = p[pcursor:pendcursor]
            print(residues)

            pcursor = p.find('Unique protein chains: ', pcursor) + 23
            pendcursor = p.find('<', pcursor)
            chains = p[pcursor:pendcursor]
            print(chains)

            # check OPM
            opm = getOPM(pdb)
            print('OPM = ' + opm)

            # OUTPUT
            details = [pdb, length, score, expect, method, identities, positives, gaps, title, organism, deposited,
                           global_stoichiometry, pseudo_stoichiometry, weight, atoms, residues, chains, opm]
            output.append(details)
            print(details)

        value = -1
        for dheader in detailheaders:
            value += 1
            print(dheader)
            print(value)
            print(details[value])
            print(all_data[row][headers['TOP_PDB'] +1 + value])
            all_data[row][headers[dheader]] = details[value]

    print ('pdbs = ' + str(pdbs))
    return pdbs

def pdb2stoichiometry(pdb):
    print('pdb2stoichiometry ' + pdb)
    au = ba1 = ''
    pdbfile = urllib2.urlopen('https://www.rcsb.org/structure/' + pdb)
    p = pdbfile.read()
    sections = p.split('BiologicalUnit')
    for section in sections:
        if "Asymmetric Unit" in section:
            au = (section.split("Global Stoichiometry</strong>: ")[1]).split('&nbsp')[0] + ' ' + (section.split('&nbsp<span style="word-wrap:break-word;">')[1]).split('&nbsp')[0]
            # print('AU = ' + au)
        if "Biological Assembly 1" in section:
            ba1 = (section.split("Global Stoichiometry</strong>: ")[1]).split('&nbsp')[0] + ' ' + (section.split('&nbsp<span style="word-wrap:break-word;">')[1]).split('&nbsp')[0]
            # print('BU1 = ' + ba1)
    if ba1:
        return 'BA1:' + ba1
    else:
        return 'AU:' + au


# Can also be used instead of Entrez package:
# urllibrary = 'http://www.ncbi.nlm.nih.gov/protein/ADH21625.1?report=fasta&log$=seqview&format=Excel#'
# for row in range(1):
#     response = urllib.urlopen(urllibrary)
#     print response.read()
#

def insertcolumn(name, position, value=''):  # be careful - inserting columns can throw off cell references! e.g. for formulae
    print('inserting column: ' + name + ' in position: ' + str(position))
    for x in headers:
        if headers[x] >= position:
            headers[x] += 1
    ws.insert_cols(position + 1)
    for y in range(len(all_data)):  # inserting cell for every row
        all_data[y].insert(position, value)
        ws.cell(position + 1, position + 1, value)
    all_data[headersrow][position] = name
    headers[name]=position
    newcolumns = 1


def checkOPM(pdbId):
    print('checkOPM ' + pdbId)
    # if not os.path.isfile(pdbpath + pdbId + "_mb.pdb"):
    search_url = "http://opm.phar.umich.edu/protein.php?search=" + pdbId  # 1l7v
    try:
        response = urllib2.urlopen(search_url)
    except:
        print ("   problem accessing " + search_url)
        return "", pdbId
    res = response.read()
    # <a href="pdb/1l7v.pdb">Download Coordinates</a>
    p = MyHTMLParser()
    p.extract_p_contents(res, lookforData="Download Coordinates", lookforTag="a", lookforAttr=['href'])
    if not len(p.stored):
        # check for reference
        pattern = "protein.php?pdbid="
        ind = res.find(pattern)
        if ind != -1:
            # do it again on new ref pdb
            newpdb = res[ind + len(pattern):ind + len(pattern) + 4]
            return checkOPM(newpdb)
        else:
            print "   problem ", pdbId
            return ""
    url = "http://opm.phar.umich.edu/" + p.stored[0][0]
    print url
    # direct access
    try:
        response = urllib2.urlopen(url)
        res = response.read()
        # f = open(model_dir + os.sep + 'PDB' + os.sep + pdbId + "_mb.pdb", "w")
        # f.write(res)
        # f.close()
        return pdbId + "_mb"
    except:
        # not found
        # build
        print ("   problem accessing " + url)
        return "", pdbId


def getOPM(pdbId):
    print('getOPM ' + pdbId)
    if not os.path.isfile(pdbpath + pdbId + "_mb.pdb"):
        search_url = "http://opm.phar.umich.edu/protein.php?search=" + pdbId  # 1l7v
        try:
            response = urllib2.urlopen(search_url)
        except:
            print ("   problem accessing " + search_url)
            return "", pdbId
        res = response.read()
        # <a href="pdb/1l7v.pdb">Download Coordinates</a>
        p = MyHTMLParser()
        p.extract_p_contents(res, lookforData="Download Coordinates", lookforTag="a", lookforAttr=['href'])
        if not len(p.stored):
            # check for reference
            pattern = "protein.php?pdbid="
            ind = res.find(pattern)
            if ind != -1:
                # do it again on new ref pdb
                newpdb = res[ind + len(pattern):ind + len(pattern) + 4]
                return getOPM(newpdb)
            else:
                print "   problem ", pdbId
                return ""
        url = "http://opm.phar.umich.edu/" + p.stored[0][0]
        print url
        # direct access
        try:
            response = urllib2.urlopen(url)
            res = response.read()
            f = open(model_dir + os.sep + 'PDB' + os.sep + pdbId + "_mb.pdb", "w")
            f.write(res)
            f.close()
            return pdbId + "_mb"
        except:
            # not found
            # build
            print ("   problem accessing " + url)
            return "", pdbId
    else:
        return pdbId + "_mb"  # mb is what Ludo uses to indicate membrane proteins - change to OPM?


def computeOPM(pdbId):
    print('computeOPM ' + pdbId)
    filename = pdbpath + pdbId + ".pdb"
    print(filename)
    if not os.path.isfile(pdbpath + pdbId + "_opmcomp_mb.pdb"):
        print('x')
        start = time.time()
        url = "http://sunshine.phar.umich.edu/upload_file.php"
        payload = {"submit":"Submit","inout":"in","yesno":"no"}#,"userfile":filename}
        print('x')
        r = requests.post(url, data=payload,files={"userfile":open(filename,"r")},timeout=600.0)
        p = MyHTMLParser()
        p.extract_p_contents(r.content,lookforData=None, lookforTag="a",lookforAttr=['href'])
        print('x')
        if not len(p.stored) :
            return ""
        url_download = "http://sunshine.phar.umich.edu/"+p.stored[0][0]
        response = urllib2.urlopen(url_download)
        print('x')
        res = response.read()
        f = open(pdbpath + pdbId + "_opmcomp_mb.pdb","w")
        f.write(res)
        f.close()
        print('x')
        end = time.time()
        print('time elapsed: ' + str(end - start))
    return pdbId + "_opmcomp_mb"


def cleanOPM(opmfile):  # gets rid of membrane atoms from OPM and finds offset
    print('cleanOPM ' + opmfile)
    opmdata = open(pdbpath + opmfile + '.pdb', "r+")
    opmdata = opmdata.read()
    lines = opmdata.split('\n')
    lastline = 1
    zmax = -10000  # proteins in OPM files might be between z values of e.g. 15 and 50
    zmin = 10000  # need to start with offset zmax and zmin to account for this
    for x in range(0,len(lines)):
        if lines[x][0:4] == 'ATOM' or lines[x][0:6] == 'HETATM':
            if lines[x][17:20] == 'DUM':
                # inner = float(lines[x][47:54])  # going to assume that membrane center is at z=0
                # outer = float(lines[x+1][47:54])
                # mcenter = (inner + outer)/2
                lastline = x
                break
            if lines[x][17:20] != 'DUM':
                if float(lines[x][46:54]) > zmax:
                    zmax = float(lines[x][46:54])
                if float(lines[x][46:54]) < zmin:
                    zmin = float(lines[x][46:54])
    offset = zmax - ((zmax - zmin) / 2)
    newfile = '\n'.join(lines[:lastline])
    cleanopm = open(model_dir + os.sep + 'PDB' + os.sep + opmfile + "_c.pdb", "w")
    print('writing ' + opmfile + '_c.pdb')
    cleanopm.write(newfile)
    cleanopm.close()
    return opmfile + '_c', offset


def calculateOffset(opmfile):  # gets rid of membrane atoms from OPM and finds offset
    print('calculateOffset ' + opmfile)
    opmdata = open(pdbpath + opmfile + '.pdb', "r+")
    opmdata = opmdata.read()
    lines = opmdata.split('\n')
    # lastline = 1
    zmax = -10000  # proteins in OPM files might be between z values of e.g. 15 and 50
    zmin = 10000  # need to start with offset zmax and zmin to account for this
    for x in range(0,len(lines)):
        if lines[x][0:4] == 'ATOM' or lines[x][0:6] == 'HETATM':
            if lines[x][17:20] == 'DUM':
                # inner = float(lines[x][47:54])  # going to assume that membrane center is at z=0
                # outer = float(lines[x+1][47:54])
                # mcenter = (inner + outer)/2
                # lastline = x
                break
            if lines[x][17:20] != 'DUM':
                if float(lines[x][46:54]) > zmax:
                    zmax = float(lines[x][46:54])
                if float(lines[x][46:54]) < zmin:
                    zmin = float(lines[x][46:54])
    offset = zmax - ((zmax - zmin) / 2)
    # newfile = '\n'.join(lines[:lastline])
    # cleanopm = open(model_dir + os.sep + 'PDB' + os.sep + opmfile + "_c.pdb", "w")
    # print('writing ' + opmfile + '_c.pdb')
    # cleanopm.write(newfile)
    # cleanopm.close()
    return offset


def write_pdbFile(pdbid):
    print('write_pdbFile (' + pdbid + ')')
    data = fetch_pdb(pdbid)
    if data[0] is not '<':  # 20170922 If it returns a legit pdb file, it starts with text. Errors start with html markup.
        pdbfile = open(pdbpath + str(pdbid) + '.pdb', 'w')
        print('writing ' + pdbid + '.pdb')
        pdbfile.write(data)
        pdbfile.close()
        return True
    else:
        return False


# given PDB ID returns PDB file information - this code was written by Jared Truong
def fetch_pdb(pdbid):
    print('fetch_pdb' + pdbid)
    url = 'http://www.rcsb.org/pdb/files/%s.pdb' % pdbid
    return urllib.urlopen(url).read()

output = []
for x in range(0,headersrow+1):
    output.append(all_data[x])


def makex(source,target):
    print('x')

def possible(x, source, target):
    if source in headers and target in headers and all_data[x][headers[source]] and not all_data[x][headers[target]]:
        return True
    else:
        return False

def main(gi=0, accession=0, ipi=0, uniparc=0, sequence=0, modifiedSequence=0, foundpdbs=0, pdb='', opm=0, cleanopm=0, temp_filename=0, localization=0, PRINCIPAL_VECTOR=0, offset=0, JITTER_MAX=0, LENGTH=0, mw=0, cofactors=0, ligands=0, interactions =0):
    proteinNumber = 1
    with open(str(model_dir + csvname + '_inprogress.csv'), 'wb') as csvprogress:  # opens progress file
        for row in range(0, headersrow+1):
            spamwriter = csv.writer(csvprogress)
            spamwriter.writerow(all_data[row])
    for x in range(headersrow + 1,len(all_data)):
        # if 'INCLUDE' in headers and all_data[x][headers['INCLUDE']] and all_data[x][headers['INCLUDE']] == 'X' or all_data[x][headers['INCLUDE']] == 'x' or all_data[x][headers['INCLUDE']] == 'TRUE':
        if 'INCLUDE' in headers and all_data[x][headers['INCLUDE']] == 'X':
            print('\nrow ' + str(x + 1))
            # uniparc = all_data[x][headers['UNIPARC']]
            # sequence = all_data[x][headers['SEQUENCE_ORIGINAL']]
            # modifiedSequence = all_data[x][headers['MODIFIED SEQUENCE']]
            # foundpdbs = all_data[x][headers['FOUNDPDBS']]
            # pdb = all_data[x][headers['chosenPDB']]
            # opm = all_data[x][headers['OPM']]
            # cleanopm = all_data[x][headers['CLEANOPM']]
            uniprot = ipi = gi = ''
            # temp_filename = ''
            # localization = all_data[x][headers['LOCALIZATION']]
            # accession = all_data[x][headers['ACCESSION']]
            # PRINCIPAL_VECTOR = all_data[x][headers['PRINCIPAL_VECTOR']]
            # offset = all_data[x][headers['OFFSET']]
            # JITTER_MAX = all_data[x][headers['JITTER_MAX']]
            # LENGTH = all_data[x][headers['LENGTH']]
            # mw = all_data[x][headers['MW']]
            # cofactors = all_data[x][headers['COFACTORS']]
            # ligands = all_data[x][headers['LIGANDS']]
            # interactions = all_data[x][headers['INTERACTIONS']]
            if 'UNIPROT_ID' in headers:
                uniprot = all_data[x][headers['UNIPROT_ID']]
            if 'UNIPROT_ID' in headers and all_data[x][headers['UNIPROT_ID']] == '' and 'UNIPROT_given' in headers:
                uniprot = all_data[x][headers['UNIPROT_given']].split(';')[0]
                all_data[x][headers['UNIPROT_ID']] = uniprot
            if uniprot:
                print(uniprot)
            if 'IPI' in headers and all_data[x][headers['IPI']]:
                ipi = all_data[x][headers['IPI']]
                ipis = ipi.split(";")
                ipi = ipis[0]
                ipi = ipi.strip()
            if 'GI' in headers and all_data[x][headers['GI']]:
                gi = all_data[x][headers['GI']]
                gi = gi.strip()
            if gi and 'ACCESSION' in headers and not all_data[x][headers['ACCESSION']]:
                try:
                    accession = gi2acc(gi)
                    all_data[x][headers['ACCESSION']] = accession
                    print('accession = ' + accession)
                except:
                    print('could not find gi')
            if 'ACCESSION' in headers:
                if all_data[x][headers['ACCESSION']] and 'SEQUENCE_ORIGINAL' in headers and not all_data[x][headers['SEQUENCE_ORIGINAL']]:
                    try:
                        sequence = acc2seq(all_data[x][headers['ACCESSION']])
                        all_data[x][headers['SEQUENCE_ORIGINAL']] = sequence
                        print('sequence = ' + sequence)
                    except:
                        print('could not find sequence')
            # if all_data[x][headers['ACCESSION']] and 'UNIPARC' in headers and not all_data[x][headers['UNIPARC']]:
            #     try:
            #         print(accession)
            #         accession = all_data[x][headers['ACCESSION']]
            #         uniparc = acc2uniparc(accession)
            #         all_data[x][headers['UNIPARC']] = uniparc
            #         print('uniparc = ' + uniparc)
            #     except:
            #         print('could not find uniparc')
            if 'ACCESSION' in headers and all_data[x][headers['ACCESSION']] and 'UNIPROT_ID' in headers and not all_data[x][headers['UNIPROT_ID']]:
                try:
                    print(accession)
                    accession = all_data[x][headers['ACCESSION']]
                    uniprot = acc2uniprot(accession)
                    all_data[x][headers['UNIPROT_ID']] = uniprot
                    print('uniprot = ' + uniprot)
                except:
                    print('could not find uniparc')
            # if all_data[x][headers['IPI']] and 'UNIPROT_ID' in headers and not all_data[x][headers['UNIPROT_ID']]:
            #     try:
            #         uniprot = ipi2uniprot(ipi)
            #         all_data[x][headers['UNIPROT_ID']] = uniprot
            #         print('uniprot = ' + uniprot)
            #     except:
            #         print('could not find uniprot')
# possibly download Uniprot data txt here? then pass to future functions? actually, already being downloaded here...
#             if uniprot:
#                 r = urllib.urlopen("http://www.uniprot.org/uniprot/" + uniprot + ".xml").read()
#                 print(r)
#                 if r:
#                     print('Uniprot is valid')
#                 else:
#                     print('Uniprot is not valid')
#                     uniprot = 'complicated'
#             if 'IPI' in headers and all_data[x][headers['IPI']] and 'UNIPARC' in headers and not all_data[x][headers['UNIPARC']] and 'UNIPROT_ID' in headers and not all_data[x][headers['UNIPROT_ID']]:
#                 ipis = all_data[x][headers['IPI']].split(';')
#                 for ipi in ipis:
#                     try:
#                         ipi = ipi.strip()
#                         uniparc, uniprot = ipi2uniparcAndUniprot(ipi)
#                     except:
#                         print('could not find uniparc')
#                     if uniparc != '':
#                         all_data[x][headers['UNIPARC']] = uniparc
#                         all_data[x][headers['UNIPROT_ID']] = uniprot
#                         break
#                 print('uniparc = ' + uniparc)
#                 print('uniprot = ' + uniprot)
#             print(all_data[x][headers['UNIPARC']])
            if 'IPI' in headers and all_data[x][headers['IPI']] and 'UNIPARC' in headers and not all_data[x][headers['UNIPARC']]:
                ipis = all_data[x][headers['IPI']].split(';')
                for ipi in ipis:
                    try:
                        ipi = ipi.strip()
                        uniparc = ipi2uniparc(ipi)
                    except:
                        print('could not find uniparc')
                    if uniparc != '':
                        all_data[x][headers['UNIPARC']] = uniparc
                        break
                print('uniparc = ' + uniparc)
#             if 'IPI' in headers and all_data[x][headers['IPI']] and 'UNIPARC' in headers and not all_data[x][headers['UNIPARC']]:
#                 try:
#                     uniparc = ipi2uniparc(all_data[x][headers['IPI']])
#                     all_data[x][headers['UNIPARC']] = uniparc
#                     print('uniparc = ' + uniparc)
#                 except:
#                     print('could not find uniparc')
#             print(all_data[x][headers['UNIPARC']])

            if 'UNIPARC' in headers and all_data[x][headers['UNIPARC']] and 'UNIPROT_ID' in headers and not all_data[x][headers['UNIPROT_ID']]:
                try:
                    uniprot = uniparc2uniprot(all_data[x][headers['UNIPARC']])
                    all_data[x][headers['UNIPROT_ID']] = uniprot
                    print('uniprot = ' + uniprot)
                except:
                    print('could not find uniprot')
            # if uniprot != '' and uniprot != '?' and uniprot != 'complicated':
            #     if not checkOrganism(uniprot):
            #         all_data[x][headers['notes']] = 'This uniprot entry is not for a rat!'
            #         try:
            #             uniprot = ipi2uniprot(ipi)
            #             print(uniprot)
            #             all_data[x][headers['UNIPROT_ID']] = uniprot
            #         except:
            #             try:
            #                 uniprot = uniparc2uniprot(uniparc)
            #                 all_data[x][headers['UNIPROT_ID']] = uniprot
            #             except:
            #                 print('could not find uniprot')

            header = 'ORGANISM_UNIPROT'
            if header in headers:
                if uniprot and not all_data[x][headers[header]]:  # BB 20180918 this is a new way - better? Can be put into a loop!
                    try:
                        all_data[x][headers[header]] = getOrganism(uniprot)
                        print(header + ' = ' + all_data[x][headers[header]])
                    except:
                        print('could not find ' + all_data[headersrow][headers[header]])

            header = 'GENE_UNIPROT'
            if header in headers:
                if uniprot and not all_data[x][headers[header]]:
                    try:
                        all_data[x][headers[header]] = getGene(uniprot)
                        print(header + ' = ' + all_data[x][headers[header]])
                    except:
                        print('could not find ' + all_data[headersrow][headers[header]])

            header = 'UNIPROT_NAME'
            if header in headers:
                if uniprot and not all_data[x][headers[header]]:
                    try:
                        all_data[x][headers[header]] = getUniprotName(uniprot)
                        print(header + ' = ' + all_data[x][headers[header]])
                    except:
                        print('could not find ' + all_data[headersrow][headers[header]])
            # header = 'REVIEWED'
            # if header in headers:
            #     if uniprot and not all_data[x][headers[header]]:
            #         try:
            #             all_data[x][headers[header]] = getReviewed(uniprot)
            #             print(header + ' = ' + all_data[x][headers[header]])
            #         except:
            #             print('could not find ' + all_data[headersrow][headers[header]])

            header = 'BEST_UNIPROT_FROM_GENE'
            if header in headers:
                if 'GENE_UNIPROT' in headers and all_data[x][headers['GENE_UNIPROT']] and not all_data[x][headers[header]]:
#                    try:
                        all_data[x][headers[header]] = getBestUniprotFromGene(all_data[x][headers['GENE_UNIPROT']])
                        print(header + ' = ' + all_data[x][headers[header]])
#                    except:
#                        print('could not find ' + all_data[headersrow][headers[header]])

            header = 'BEST_REVIEWED'
            if header in headers:
                if 'BEST_UNIPROT_FROM_GENE' in headers and all_data[x][headers['BEST_UNIPROT_FROM_GENE']] and not all_data[x][headers[header]]:
                    try:
                        all_data[x][headers[header]] = getReviewed(all_data[x][headers['BEST_UNIPROT_FROM_GENE']])
                        print(header + ' = ' + all_data[x][headers[header]])
                    except:
                        print('could not find ' + all_data[headersrow][headers[header]])

            header = 'BEST_SCORE'
            if header in headers:
                if 'BEST_UNIPROT_FROM_GENE' in headers and all_data[x][headers['BEST_UNIPROT_FROM_GENE']] and not all_data[x][headers[header]]:
                    try:
                        all_data[x][headers[header]] = getAnnotationScore(all_data[x][headers['BEST_UNIPROT_FROM_GENE']])
                        print(header + ' = ' + all_data[x][headers[header]])
                    except:
                        print('could not find ' + all_data[headersrow][headers[header]])

            # for entry in ['ORGANISM']:
            #     if all_data[x][headers['UNIPROT_ID']] and not all_data[x][headers[entry]]:  # BB 20180918 this is a new way - better? Can be put into a loop!
            #         try:
            #             all_data[x][headers[entry]] = getUniprotInfo(uniprot)
            #             print(all_data[headersrow][headers[entry]] + ' = ' + all_data[x][headers[entry]])
            #         except:
            #             print('could not find ' + all_data[headersrow][headers[entry]])
            # if uniprot and not all_data[x][headers['ORGANISM']]:
            #     try:
            #         localization = uniprot2localization(uniprot)
            #         all_data[x][headers['LOCALIZATION']] = localization
            #         print('localization = ' + localization)
            #     except:
            #         print('could not find localization')
            # if uniprot and not localization:
            #     try:
            #         localization = uniprot2localization(uniprot)
            #         all_data[x][headers['LOCALIZATION']] = localization
            #         print('localization = ' + localization)
            #     except:
            #         print('could not find localization')

            header = 'UNIPROT_SCORE'
            if header in headers:
                if uniprot and not all_data[x][headers[header]]:
                    try:
                        all_data[x][headers[header]] = getAnnotationScore(uniprot)
                        print(header + ' = ' + all_data[x][headers[header]])
                    except:
                        print('could not find ' + all_data[headersrow][headers[header]])


            # These routines are a way to find the best Uniprot entry, but I'm not convinced they're worthwhile. Superceded by getBestUnipotFromGene.
            header = 'UNIREF50'
            if header in headers:
                if uniprot and not all_data[x][headers[header]]:
                    try:
                        all_data[x][headers[header]] = getUniref50(uniprot)
                        print(header + ' = ' + all_data[x][headers[header]])
                    except:
                        print('could not find ' + all_data[headersrow][headers[header]])
            header = 'UNIREF90'
            if header in headers:
                if uniprot and not all_data[x][headers[header]]:
                    try:
                        all_data[x][headers[header]] = getUniref90(uniprot)
                        print(header + ' = ' + all_data[x][headers[header]])
                    except:
                        print('could not find ' + all_data[headersrow][headers[header]])
            header = 'UNIREF100'
            if header in headers:
                if uniprot and not all_data[x][headers[header]]:
                    try:
                        all_data[x][headers[header]] = getUniref100(uniprot)
                        print(header + ' = ' + all_data[x][headers[header]])
                    except:
                        print('could not find ' + all_data[headersrow][headers[header]])
            header = 'UNIREF50_TOP'
            if header in headers:
                if all_data[x][headers['UNIREF50']] and not all_data[x][headers[header]]:
                    try:
                        all_data[x][headers[header]] = getUniRef50Top(all_data[x][headers['UNIREF50']])
                        print(header + ' = ' + all_data[x][headers[header]])
                    except:
                        print('exception ' + all_data[headersrow][headers[header]])
            header = 'UNIREF50_SCORE'
            if header in headers:
                if all_data[x][headers['UNIREF50_TOP']] and all_data[x][headers['UNIREF50_TOP']] != 'N/A' and not all_data[x][headers[header]]:
                    try:
                        all_data[x][headers[header]] = getAnnotationScore(all_data[x][headers['UNIREF50_TOP']])
                        print(header + ' = ' + all_data[x][headers[header]])
                    except:
                        print('could not find ' + all_data[headersrow][headers[header]])
                else:
                    all_data[x][headers[header]] = 'N/A'
            header = 'UNIREF90_TOP'
            if header in headers:
                if all_data[x][headers['UNIREF90']] and not all_data[x][headers[header]]:
                    try:
                        all_data[x][headers[header]] = getUniRef90Top(all_data[x][headers['UNIREF90']])
                        print(header + ' = ' + all_data[x][headers[header]])
                    except:
                        print('exception ' + all_data[headersrow][headers[header]])
            header = 'UNIREF90_SCORE'
            if header in headers:
                if all_data[x][headers['UNIREF90_TOP']] and all_data[x][headers['UNIREF90_TOP']] != 'N/A' and not all_data[x][headers[header]]:
                    try:
                        all_data[x][headers[header]] = getAnnotationScore(all_data[x][headers['UNIREF90_TOP']])
                        print(header + ' = ' + all_data[x][headers[header]])
                    except:
                        print('could not find ' + all_data[headersrow][headers[header]])
                else:
                    all_data[x][headers[header]] = 'N/A'
            header = 'UNIREF100_TOP'
            if header in headers:
                if all_data[x][headers['UNIREF100']] and not all_data[x][headers[header]]:
                    try:
                        all_data[x][headers[header]] = getUniRef100Top(all_data[x][headers['UNIREF100']])
                        print(header + ' = ' + all_data[x][headers[header]])
                    except:
                        print('exception ' + all_data[headersrow][headers[header]])
            header = 'UNIREF100_SCORE'
            if header in headers:
                if all_data[x][headers['UNIREF100_TOP']] and all_data[x][headers['UNIREF100_TOP']] != 'N/A' and not all_data[x][headers[header]]:
                    try:
                        all_data[x][headers[header]] = getAnnotationScore(all_data[x][headers['UNIREF100_TOP']])
                        print(header + ' = ' + all_data[x][headers[header]])
                    except:
                        print('could not find ' + all_data[headersrow][headers[header]])
                else:
                    all_data[x][headers[header]] = 'N/A'

            header = 'BEST_UNIPROT'
            if header in headers:
                if all_data[x][headers['UNIREF100_TOP']] and not all_data[x][headers[header]]:
                    try:
                        all_data[x][headers[header]] = getAnnotationScore(all_data[x][headers['UNIREF100_TOP']])
                        print(header + ' = ' + all_data[x][headers[header]])
                    except:
                        print('could not find ' + all_data[headersrow][headers[header]])

            source, target = 'UNIPROT_ID', 'LOCALIZATION'
            if source in headers and target in headers and all_data[x][headers[source]] and not all_data[x][headers[target]]:
                try:
                    all_data[x][headers[target]] = uniprot2localization(all_data[x][headers[source]])
                except:
                    print('   could not get ' + target)

            # BB - I don't think this works as well as the original
            # source, target = 'UNIPROT_ID', 'LOCALIZATIONSV'
            # if source in headers and target in headers and all_data[x][headers[source]] and not all_data[x][headers[target]]:
            #     try:
            #         all_data[x][headers[target]] = uniprot2localizationSV(all_data[x][headers[source]])
            #     except:
            #         print('   could not get ' + target)
            #

            source, target = 'LOCALIZATION', 'MEMBRANE'
            if source in headers and target in headers and all_data[x][headers[source]] and not all_data[x][headers[target]]:
                if all_data[x][headers[source]].find('embrane') != -1:  # LEAVING OFF FIRST CHARACTER TO ACCOUNT FOR CAPITALS
                    all_data[x][headers['MEMBRANE']] = 'X'
                    print('SETTING MEMBRANE = X')

            # if uniparc and not sequence:
            #     try:
            #         sequence = uniparc2seq(uniparc)
            #         all_data[x][headers['SEQUENCE_ORIGINAL']] = sequence
            #     except:
            #         print('   could not get sequence')

            if uniprot and 'SEQUENCE_ORIGINAL' in headers and not all_data[x][headers['SEQUENCE_ORIGINAL']]:
                try:
                    if uniprot != '' and uniprot != '?' and uniprot != 'complicated':
                        sequence = uniprot2seq(uniprot)
                        all_data[x][headers['SEQUENCE_ORIGINAL']] = sequence
                        if sequence:
                            print("successful")
                except:
                    print(' could not get sequence')

            source1, source2, target = 'UNIPROT_ID', 'SEQUENCE_ORIGINAL', 'SEQUENCE_FINAL'
            if source1 in headers and target in headers and source2 in headers and all_data[x][headers[source1]] \
                    and all_data[x][headers[source2]] and not all_data[x][headers[target]]:
                all_data[x][headers[target]] = cleaveSignalPeptide(all_data[x][headers[source1]], all_data[x][headers[source2]])

            source, target = 'SEQUENCE_FINAL', 'FOUNDPDBS'
            if source in headers and target in headers and all_data[x][headers[source]] and not all_data[x][headers[target]]:
                # try:
                    all_data[x][headers[target]] = ';'.join(seq2pdb(x, all_data[x][headers[source]], details=1))
                # except:
                #     print('   could not get foundpdbs')


            if possible(x, 'TOP_PDB', 'PDB_STOICHIOMETRY'):
                # try:
                    all_data[x][headers['PDB_STOICHIOMETRY']] = pdb2stoichiometry(all_data[x][headers['TOP_PDB']])
                    print(all_data[x][headers['PDB_STOICHIOMETRY']])
                # except:
                #     print('   could not get PDB_STOICHIOMETRY')

            # Not sure what I was trying to do here. FIX!
            # source = 'TOP_PDB'
            # if source in headers and all_data[x][headers[source]]:
            #     try:
            #         all_data[x][headers[target]] = ';'.join(seq2pdb(all_data[x][headers[source]]))
            #     except:
            #         print('   could not get foundpdbs')

            # if foundpdbs and not pdb:
            #     pdb = foundpdbs[0]
            #     temp_filename = pdbpath + pdb + ".pdb"
            #     if not os.path.isfile(temp_filename):
            #         write_pdbFile(pdb)
            # elif pdb:
            #     temp_filename = pdbpath + pdb + ".pdb"
            #     if not os.path.isfile(temp_filename):
            #         write_pdbFile(pdb)

            source, target = 'FOUNDPDBS', 'TOP_PDB'
            if source in headers and target in headers and all_data[x][headers[source]] and not all_data[x][headers[target]]:
                try:
                    all_data[x][headers[target]] = (all_data[x][headers[source]].split(';'))[0]
                except:
                    print('   could not get ' + target)

            source, target = 'TOP_PDB', 'OPM'
            if source in headers and target in headers and all_data[x][headers[source]] and not all_data[x][headers[target]]:
                try:
                    all_data[x][headers[target]] = getOPM(all_data[x][headers[source]])
                except:
                    print('   could not get ' + target)

            source, target = 'OPM', 'MEMBRANE'
            if source in headers and target in headers and all_data[x][headers[source]] and not all_data[x][headers[target]]:
                try:
                    all_data[x][headers[target]] = 'X'
                except:
                    print('   could not get ' + target)

            # if pdb and not opm:
            #     opm = getOPM(pdb)

            source1, source2, target = 'TOP_PDB', 'MEMBRANE', 'OPM'
            if source1 in headers and target in headers and source2 in headers and all_data[x][headers[source1]] \
                    and all_data[x][headers[source2]] and not all_data[x][headers[target]]:
                try:
                    all_data[x][headers[target]] = computeOPM(all_data[x][headers[source1]])
                except:
                    print('could not compute OPM ' + all_data[x][headers[source1]])

            # if 'MEMBRANE' in headers:
            #     if opm == "" and all_data[x][headers['MEMBRANE']] != '':
            #         print('computing OPM')
            #         try:
            #             opm = computeOPM(temp_filename,pdb)
            #         except:
            #             print('could not compute OPM ' + pdb)
            # if opm and not all_data[x][headers['OPM']]:
            #     all_data[x][headers['MEMBRANE']] = 'X'
            #     all_data[x][headers['OPM']] = opm
            # if opm and not cleanopm:
            #     cleanopm, _ = cleanOPM(opm)
            #     all_data[x][headers['CLEANOPM']] = cleanopm

            source, target = 'OPM', 'CLEANOPM'
            if source in headers and target in headers and all_data[x][headers[source]] and not all_data[x][headers[target]]:
                try:
                    all_data[x][headers[target]] = cleanOPM(all_data[x][headers[source]])[0]
                except:
                    print('   could not get ' + target)

            source1, source2, target = 'CLEANOPM', 'TOP_PDB', 'PDB_FINAL'
            if source1 in headers and target in headers and all_data[x][headers[source1]] and not all_data[x][headers[target]]:
                all_data[x][headers[target]] = all_data[x][headers[source1]]
            elif source2 in headers and target in headers and all_data[x][headers[source2]] and not all_data[x][headers[target]]:
                all_data[x][headers[target]] = all_data[x][headers[source2]]

            # if possible(x, source='TOP_PDB', target='PDB_ORGANISM'):
            #     makex(source, target)
            #
            # if possible(x, source='TOP_PDB', target='PDB_ORGANISM'):
            #     try:
            #         all_data[x][headers[target]] = pdb2organism(all_data[x][headers[source]])[0]
            #     except:
            #         print('   could not get ' + target)

            # if possibledo(x, source='TOP_PDB', target='PDB_ORGANISM', makex(source, target)):

            # if opm and not offset:
            #     _, offset = cleanOPM(opm)
            #     all_data[x][headers['OFFSET']] = '0, 0, ' + str(offset)

            source, target = 'CLEANOPM', 'OFFSET'
            if source in headers and target in headers and all_data[x][headers[source]] and not all_data[x][headers[target]]:
                try:
                    all_data[x][headers[target]] = '0, 0, ' + str(calculateOffset(all_data[x][headers[source]]))
                except:
                    print('   could not get ' + target)

            # if opm and not PRINCIPAL_VECTOR:
            #     all_data[x][headers['PRINCIPAL_VECTOR']] = '0, 0, -1'  # must be negative for vesicular membrane proteins.

            source, target = 'CLEANOPM', 'PRINCIPAL_VECTOR'
            if source in headers and target in headers and all_data[x][headers[source]] and not all_data[x][headers[target]]:
                print('setting ' + target)
                try:
                    all_data[x][
                        headers[target]] = '0, 0, -1'  # must be negative for vesicular membrane proteins.
                except:
                    print('   could not get ' + target)

            # if opm and not JITTER_MAX:
            #     all_data[x][headers['JITTER_MAX']] = '0.1, 0.1, 0'

            source, target = 'CLEANOPM', 'JITTER_MAX'
            if source in headers and target in headers and all_data[x][headers[source]] and not all_data[x][headers[target]]:
                print('setting ' + target)
                try:
                    all_data[x][
                        headers[target]] = '0.1, 0.1, 0'  # must be negative for vesicular membrane proteins.
                except:
                    print('   could not get ' + target)

            if 'chosenPDB' in headers:
                if cleanopm and not all_data[x][headers['chosenPDB']]:
                    all_data[x][headers['chosenPDB']] = cleanopm

            if 'chosenPDB' in headers:
                if not all_data[x][headers['chosenPDB']]:
                    all_data[x][headers['chosenPDB']] = pdb
            # if 'PDB' in headers:
            #      if not all_data[x][headers['PDB']]:
            #         all_data[x][headers['PDB']] = pdb

            source, target = 'SEQUENCE_FINAL', 'MW'
            if possible(x, 'SEQUENCE_FINAL', 'MW'):
                try:
                    seq = ProteinAnalysis(all_data[x][headers['SEQUENCE_FINAL']])
                    mw = seq.molecular_weight()
                    all_data[x][headers['MW']] = mw
                except:
                    all_data[x][headers['MW']] = 'fail'
                    print('MW failed')
                    pass

            if possible(x, 'SEQUENCE_ORIGINAL', 'LENGTH_SEQ_ORIGINAL'):
                all_data[x][headers['LENGTH_SEQ_ORIGINAL']] = len(all_data[x][headers['SEQUENCE_ORIGINAL']])

            if possible(x, 'SEQUENCE_FINAL', 'LENGTH_SEQ_FINAL'):
                all_data[x][headers['LENGTH_SEQ_FINAL']] = len(all_data[x][headers['SEQUENCE_FINAL']])

            # if sequence and 'SEQUENCE_FINAL in headers and not modifiedSequence:
            #     modifiedSequence = cleaveSignalPeptide(uniprot, sequence)
            #     all_data[x][headers['SEQUENCE_FINAL]]  = modifiedSequence

            # This routine will make a text file suitable for batch processing with Phyre2
            if 'UNIPROT_ID' in headers and all_data[x][headers['UNIPROT_ID']] and 'SEQUENCE_FINAL' in headers:
                if proteinNumber == 1:
                    print('starting allSequences.txt')
                    f = open(str(model_dir + csvname + '_allSequences.txt'), 'w')
                else:
                    f = open(str(model_dir + csvname + '_allSequences.txt'), 'a')  # bb a="append"
                f.write('\n' + '>' + all_data[x][headers['UNIPROT_ID']])
                f.write('\n'+ all_data[x][headers['SEQUENCE_FINAL']])
                f.close()
                proteinNumber += 1
                # elif all_data[x][headers['SEQUENCE_ORIGINAL']]:
                #     if proteinNumber == 1:
                #         print('starting allSequences.txt')
                #         f = open(str(model_dir + csvname + '_allSequences.txt'), 'w')
                #     else:
                #         f = open(str(model_dir + csvname + '_allSequences.txt'), 'a')
                #     f.write('\n' + '>' + all_data[x][headers['UNIPROT_ID']])
                #     f.write('\n'+ all_data[x][headers['SEQUENCE_ORIGINAL']])
                #     f.close()
                #     proteinNumber += 1

            # These routines retrieve dependency and interaction info
            if 'UNIPROT_ID' in headers:
                uniprot = all_data[x][headers['UNIPROT_ID']]
            if possible(x, 'UNIPROT_ID','COFACTORS_UNIPROT'):
                if uniprot != '' and uniprot != '?' and uniprot != 'complicated' and uniprot != 'None':
                    if checkOrganism(uniprot):  # this whole subroutine needs to be fixed. FIX
                        urllib.urlretrieve("http://www.uniprot.org/uniprot/" + uniprot + ".xml", model_dir + 'uniprotentry' + uniprot + '.xml')
                        if possible(x, 'UNIPROT_ID', 'COFACTORS_UNIPROT'):
                            all_data[x][headers['COFACTORS_UNIPROT']] = findCofactorsAndLigands(uniprot)[0]
                        if possible(x, 'UNIPROT_ID', 'LIGANDS_UNIPROT'):
                            all_data[x][headers['LIGANDS_UNIPROT']] = findCofactorsAndLigands(uniprot)[1]
                        # listOfDependencies will include any found information from RGD, uniprot, stringdb, BioGrid, and IntAct
                        if possible(x, 'UNIPROT_ID', 'SUBUNIT_UNIPROT'):
                            try:
                                uniprotData = uniprotInteractions(uniprot)
                                all_data[x][headers['SUBUNIT_UNIPROT']] = uniprotData[0]
                                all_data[x][headers['UNIPROT_INTERACTIONS']] = uniprotData[1]
                            except:
                                print('uniprotInteractions failed')
                        if possible(x, 'UNIPROT_ID', 'RGD_INTERACTIONS'):  # possibly get rid of this - check to see if valuable
                            try:
                                all_data[x][headers['RGD_INTERACTIONS']] = rgdDependencies(uniprot)
                            except:
                                print('rgdDependencies failed')
                        if possible(x, 'UNIPROT_ID', 'STRINGDB_INTERACTIONS'):
                            try:
                                all_data[x][headers['STRINGDB_INTERACTIONS']] = stringdbInteractors(uniprot)
                            except:
                                print('stringdbInteractors failed')
                        if possible(x, 'UNIPROT_ID', 'BIOGRID_INTERACTIONS'):
                            # try:
                                all_data[x][headers['BIOGRID_INTERACTIONS']] = biogridInteractors(uniprot)
                            # except:
                            #     print('biogridInteractors failed')
                        if possible(x, 'UNIPROT_ID', 'INTACT_INTERACTIONS'):
                            # try:
                                all_data[x][headers['INTACT_INTERACTIONS']] = intactInteractors(uniprot)
                            # except:
                            #     print('intactInteractors failed')
                        # print(listOfDependencies)

            if 'ALL_INTERACTIONS' in headers and not all_data[x][headers['ALL_INTERACTIONS']]:
                print('adding ALL_INTERACTIONS')
                allInteractions = ''
                header = 'UNIPROT_INTERACTIONS'
                if header in headers:
                    allInteractions += all_data[x][headers[header]]

                header = 'STRINGDB_INTERACTIONS'
                if header in headers:
                    allInteractions += all_data[x][headers[header]]

                header = 'BIOGRID_INTERACTIONS'
                if header in headers:
                    allInteractions += all_data[x][headers[header]]

                header = 'INTACT_INTERACTIONS'
                if header in headers:
                    allInteractions += all_data[x][headers[header]]

                allInteractionsLower = allInteractions.lower()  # puts all in lower case

                all_data[x][headers['ALL_INTERACTIONS']] = ' '.join(set(allInteractionsLower.split(' ')))  # gets rid of duplicates

            if possible(x, 'ALL_INTERACTIONS', 'INTERACTION_COMPARTMENTS'):
                interactions = all_data[x][headers['ALL_INTERACTIONS']].split(' ')
                for interaction in interactions:
                    if interaction in interactionDict:
                        all_data[x][headers['INTERACTION_COMPARTMENTS']] += interactionDict[interaction] + ';'
                    else:
                        uniprot = getBestUniprotFromGene(interaction)
                        localization = uniprot2localization(uniprot)
                        if localization != '':
                            all_data[x][headers['INTERACTION_COMPARTMENTS']] += localization + ';'
                        interactionDict[interaction]=localization

            if possible(x, 'INTERACTION_COMPARTMENTS', 'INTERACTION_SCORES'):
                print('tallying interaction scores')
                compartments = all_data[x][headers['INTERACTION_COMPARTMENTS']].split(';')
                target=neutral=contaminant=0
                for entry in compartments:
                    if entry == '':
                        continue
                    elif entry in targetLocations:
                        target += 1
                    elif entry in neutralLocations:
                        neutral += 1
                    elif entry in contaminantLocations:
                        contaminant += 1
                    elif entry not in newLocations:
                        newLocations.append(entry)
                all_data[x][headers['INTERACTION_SCORES']] = ';'.join([str(target), str(neutral), str(contaminant)])

            if possible(x, 'INTERACTION_SCORES', 'INTERACTION_MULTIPLIER'):
                print('tallying interaction multiplier')
                scores = all_data[x][headers['INTERACTION_SCORES']].split(';')
                all_data[x][headers['INTERACTION_MULTIPLIER']] = 1.1**int(scores[0]) * 0.9**int(scores[2])
            elif 'INTERACTION_MULTIPLIER' in headers and not all_data[x][headers['INTERACTION_MULTIPLIER']]:
                all_data[x][headers['INTERACTION_MULTIPLIER']] = 1

            # The following routine will get scores from the 'LOCALIZATION'
            if possible(x, 'LOCALIZATION', 'LOCALIZATION_SCORE'):
                print('tallying localization scores')
                compartments = all_data[x][headers['LOCALIZATION']].split(';')
                target=neutral=contaminant=0
                for entry in compartments:
                    if entry == '':
                        continue
                    elif entry in targetLocalizations:
                        target += 1
                    elif entry in neutralLocalizations:
                        neutral += 1
                    elif entry in contaminantLocalizations:
                        contaminant += 1
                    elif entry not in newLocations:
                        newLocations.append(entry)
                all_data[x][headers['LOCALIZATION_SCORE']] = ';'.join([str(target), str(neutral), str(contaminant)])

            # Generate LOCALIZATION_MULTIPLIER from LOCALIZATION_SCORE - localization in neutral locations is penalized
            if possible(x, 'LOCALIZATION_SCORE', 'LOCALIZATION_MULTIPLIER'):
                print('tallying LOCALIZATION_MULTIPLIER')
                scores = all_data[x][headers['LOCALIZATION_SCORE']].split(';')
                all_data[x][headers['LOCALIZATION_MULTIPLIER']] = 1.1**int(scores[0]) * 0.9**int(scores[2])
            elif 'LOCALIZATION_MULTIPLIER' in headers and not all_data[x][headers['LOCALIZATION_MULTIPLIER']]:
                all_data[x][headers['LOCALIZATION_MULTIPLIER']] = 1

            # Take LOCALIZATION - if there is column, increment 1, if not, make column, increment 1.
            if possible(x, 'LOCALIZATION', 'FORTRAN_LOCALIZATION'):
                print('FORTRAN_LOCALIZATION computations')
                compartments = all_data[x][headers['LOCALIZATION']].split(';')
                print(compartments)
                for entry in compartments:
                    if entry != '':
                        entry = "L_" + entry
                        print(entry)
                        if entry in headers and all_data[x][headers[entry]]=='':
                            all_data[x][headers[entry]] = 1
                        elif entry in headers:
                            all_data[x][headers[entry]] += 1
                        else:
                            insertcolumn(entry, headers['FORTRAN_LOCALIZATION']+1, value=0)
                            all_data[x][headers[entry]] = 1

            # Take LOCALIZATION - if there is column, increment 1, if not, make column, increment 1.
            if possible(x, 'LOCALIZATION', 'FORTRAN_INTERACTIONS'):
                print('FORTRAN_INTERACTIONS computations')
                compartments = all_data[x][headers['INTERACTION_COMPARTMENTS']].split(';')
                print(compartments)
                for entry in compartments:
                    if entry != '':
                        entry = "I_" + entry
                        print(entry)
                        if entry in headers and all_data[x][headers[entry]]=='':
                            all_data[x][headers[entry]] = 1
                        elif entry in headers:
                            all_data[x][headers[entry]] += 1
                        else:
                            insertcolumn(entry, headers['FORTRAN_INTERACTIONS']+1, value=0)
                            all_data[x][headers[entry]] = 1

        output.append(all_data[x])

        # write entire results as _inprogress file so if crashes, everything saved. Can't just append lines, because won't work if columns are added.
        with open(str(model_dir + csvname + '_inprogress.csv'), 'wb') as csvprogress:  # writes _inprogress file
            spamwriter = csv.writer(csvprogress)
            for row in range(len(output)):
                spamwriter.writerow(output[row])
            # print('appended row ' + str(x+1) + ' to _inprogress file')

        # save interactionDict, because this represents a lot of work, and if saved, don't have to do again.
        with open(str(model_dir + csvname + '_' + str(taxid) + '_interactionDictionary.txt'), 'w') as intDict:  # saving interactionDict
            intDict.write(str(interactionDict))

main()

os.rename(model_dir + csvname + '_inprogress.csv', model_dir + csvname + '_complete.csv')

# print('interactionDict = ' + str(interactionDict))
print('newLocations = ' + str(newLocations))
print('interactionDict = ' + str(interactionDict))

# if csvpath.endswith('.xls') or csvpath.endswith('.xlsx'):
#     from openpyxl.styles import Font
#     from openpyxl.styles import Color
#     from openpyxl.styles import colors
#     red = Font(color=colors.RED)  # provides text for font
#     rowNumber = -1
#     for row in ws.rows:
#         rowNumber += 1
#         columnNumber = -1
#         for cell in row:
#             columnNumber += 1
#             if cell.value != None and cell.value != all_data[rowNumber][columnNumber]:
#                 cell.font = red
#                 cell.value = all_data[rowNumber][columnNumber]
# #     a1 = ws['A1']  # assign a variable for a cell
# #     ws.insert_cols(idx=6, amount=1)  # insert one column before column 6 - can't use letters
# #     a1.value = 'new value in red'
# #     ws['a2'].value = 45  # change value
#     wb.save(str(model_dir + csvname + '_complete.xlsx'))  # save


if csvpath.endswith('.xls') or csvpath.endswith('.xlsx'):
    from openpyxl.styles import Font
    from openpyxl.styles import Color
    from openpyxl.styles import colors
    red = Font(color=colors.RED)  # provides text for font
    for row in range(len(all_data)):
        for column in range(len(all_data[row])):
            if all_data[row][column] != '' and all_data[row][column] != ws.cell(row+1,column+1).value:
                ws.cell(row + 1, column + 1).font = red
                ws.cell(row + 1, column + 1).value = all_data[row][column]
#     a1 = ws['A1']  # assign a variable for a cell
#     ws.insert_cols(idx=6, amount=1)  # insert one column before column 6 - can't use letters
#     a1.value = 'new value in red'
#     ws['a2'].value = 45  # change value
    wb.save(str(model_dir + csvname + '_complete.xlsx'))  # save

# if csvpath.endswith('.xls') or csvpath.endswith('.xlsx'):
#     from openpyxl.styles import Font
#     from openpyxl.styles import Color
#     from openpyxl.styles import colors
#     red = Font(color=colors.RED)  # provides text for font
#     for row in range(len(output)):
#         for column in range(len(output[row])):
#             if output[row][column] != '' and output[row][column] != ws.cell(row+1,column+1).value:
#                 ws.cell(row + 1, column + 1).font = red
#                 ws.cell(row + 1, column + 1).value = output[row][column]
# #     a1 = ws['A1']  # assign a variable for a cell
# #     ws.insert_cols(idx=6, amount=1)  # insert one column before column 6 - can't use letters
# #     a1.value = 'new value in red'
# #     ws['a2'].value = 45  # change value
#     wb.save(str(model_dir + csvname + '_complete_output.xlsx'))  # save

if newcolumns == 1:
    print('new columns have been added')
print("done")
endtime = time.time()
print("time elapsed = " + str(endtime - starttime))
